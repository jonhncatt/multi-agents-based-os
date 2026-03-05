from __future__ import annotations

import json
import fnmatch
import hashlib
import re
import shlex
import shutil
import ssl
import subprocess
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from html import unescape
from pathlib import Path
from typing import Any

from app.config import AppConfig
from app.document_text import (
    build_pdf_document_index,
    clear_pdf_cache_for_path,
    extract_heading_entries_from_pages,
    extract_pdf_page_texts_from_path,
    extract_pdf_tables_from_path,
    extract_pdf_text_from_bytes,
    extract_pdf_text_from_path,
    normalize_lookup_text,
    truncate_text,
)
from app.sandbox import DockerSandboxManager


def _is_within(path: Path, root: Path) -> bool:
    return path == root or root in path.parents


def _build_path_candidates(config: AppConfig, raw_path: str) -> list[Path]:
    raw = (raw_path or ".").strip() or "."
    path = Path(raw).expanduser()
    seen: set[str] = set()
    candidates: list[Path] = []

    def add(p: Path) -> None:
        resolved = p.resolve()
        key = str(resolved)
        if key in seen:
            return
        seen.add(key)
        candidates.append(resolved)

    if path.is_absolute():
        add(path)
        return candidates

    normalized = raw.replace("\\", "/").strip("/").lower()
    normalized_slash = raw.replace("\\", "/").strip("/")
    if normalized:
        # High-priority alias mapping, e.g. "workbench/a.txt" -> "<allowed_root_named_workbench>/a.txt"
        for root in config.allowed_roots:
            root_norm = str(root).replace("\\", "/").rstrip("/").lower()
            if normalized == root_norm or normalized == root.name.lower():
                add(root)
                continue
            prefix = f"{root.name.lower()}/"
            if normalized.startswith(prefix):
                suffix = normalized_slash[len(prefix) :]
                add(root / suffix)

    # Default mapping keeps backward compatibility.
    add(config.workspace_root / path)
    for root in config.allowed_roots:
        if root == config.workspace_root:
            continue
        add(root / path)

    return candidates


def _resolve_workspace_path(config: AppConfig, raw_path: str) -> Path:
    if config.allow_any_path:
        path = Path((raw_path or ".").strip() or ".").expanduser()
        if not path.is_absolute():
            path = config.workspace_root / path
        path = path.resolve()
        return path

    candidates = _build_path_candidates(config, raw_path)

    # Prefer existing paths in allowed roots for better UX with relative inputs.
    for path in candidates:
        for root in config.allowed_roots:
            if _is_within(path, root) and path.exists():
                return path

    # Fall back to first allowed candidate even if it does not exist,
    # prefer a candidate whose parent directory exists.
    for path in candidates:
        for root in config.allowed_roots:
            if _is_within(path, root) and path.parent.exists():
                return path

    # Last resort: return first allowed candidate even if parent does not exist,
    # so upper layers can return a clear "not found" error.
    for root in config.allowed_roots:
        for path in candidates:
            if _is_within(path, root):
                return path

    allowed = ", ".join(str(p) for p in config.allowed_roots)
    raise ValueError(f"Path out of allowed roots: {raw_path}. Allowed roots: {allowed}")


def _resolve_source_path(config: AppConfig, raw_path: str) -> Path:
    """
    Resolve existing source file path with upload-name fallback.
    If raw_path is only an original upload filename (e.g. a.zip),
    try matching uploads_dir entry like <uuid>__a.zip.
    """
    resolved = _resolve_workspace_path(config, raw_path)
    if resolved.exists():
        return resolved

    raw = (raw_path or "").strip()
    if not raw:
        return resolved

    p = Path(raw.replace("\\", "/"))
    if p.is_absolute():
        return resolved

    basename = p.name
    if not basename:
        return resolved

    try:
        matches = sorted(
            config.uploads_dir.glob(f"*__{basename}"),
            key=lambda m: m.stat().st_mtime,
            reverse=True,
        )
    except Exception:
        return resolved

    for match in matches:
        candidate = match.resolve()
        for root in config.allowed_roots:
            if _is_within(candidate, root):
                return candidate
    return resolved


def _truncate_output(text: str, max_chars: int = 12000) -> str:
    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars]}\n\n[output truncated: {len(text)} chars]"


def _looks_like_html(content_type: str, text: str) -> bool:
    lower_ct = (content_type or "").lower()
    if "text/html" in lower_ct or "application/xhtml+xml" in lower_ct:
        return True
    head = text[:400].lower()
    return "<html" in head or "<!doctype html" in head


def _extract_html_text(raw_html: str, max_chars: int) -> str:
    html = re.sub(r"(?is)<!--.*?-->", " ", raw_html)
    html = re.sub(r"(?is)<(script|style|noscript).*?>.*?</\1>", " ", html)
    html = re.sub(r"(?i)<br\s*/?>", "\n", html)
    html = re.sub(r"(?i)</(p|div|li|tr|h1|h2|h3|h4|h5|h6|section|article)>", "\n", html)
    html = re.sub(r"(?s)<[^>]+>", " ", html)
    html = unescape(html)

    lines: list[str] = []
    for line in html.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip()
        if normalized:
            lines.append(normalized)

    out = "\n".join(lines)
    if len(out) > max_chars:
        out = out[:max_chars]
    return out


def _find_html_meta_content(raw_html: str, attr_name: str, attr_value: str) -> str:
    pattern = re.compile(
        rf'(?is)<meta[^>]*{attr_name}\s*=\s*["\']{re.escape(attr_value)}["\'][^>]*content\s*=\s*["\'](.*?)["\']'
    )
    match = pattern.search(raw_html or "")
    if match:
        return _clean_html_fragment(match.group(1) or "")
    return ""


def _extract_html_metadata(raw_html: str, base_url: str = "") -> dict[str, str]:
    html = raw_html or ""
    title = ""
    title_match = re.search(r"(?is)<title[^>]*>(.*?)</title>", html)
    if title_match:
        title = _clean_html_fragment(title_match.group(1) or "")
    if not title:
        title = _find_html_meta_content(html, "property", "og:title") or _find_html_meta_content(html, "name", "twitter:title")

    published_at = (
        _find_html_meta_content(html, "property", "article:published_time")
        or _find_html_meta_content(html, "name", "article:published_time")
        or _find_html_meta_content(html, "property", "og:updated_time")
        or _find_html_meta_content(html, "name", "pubdate")
        or _find_html_meta_content(html, "name", "publish-date")
    )

    canonical_url = ""
    canonical_match = re.search(r'(?is)<link[^>]*rel\s*=\s*["\']canonical["\'][^>]*href\s*=\s*["\'](.*?)["\']', html)
    if canonical_match:
        canonical_url = unescape(canonical_match.group(1) or "").strip()
        if canonical_url and base_url:
            canonical_url = urllib.parse.urljoin(base_url, canonical_url)

    return {
        "title": title,
        "published_at": published_at,
        "canonical_url": canonical_url,
    }


def _tokenize_query(query: str) -> list[str]:
    text = (query or "").strip()
    if not text:
        return []

    ascii_stopwords = {
        "a",
        "an",
        "and",
        "are",
        "at",
        "for",
        "from",
        "how",
        "in",
        "is",
        "it",
        "latest",
        "mlb",
        "news",
        "npb",
        "of",
        "on",
        "recent",
        "score",
        "scores",
        "the",
        "today",
        "what",
        "when",
        "where",
        "who",
        "why",
    }
    cjk_fillers = (
        "查一下",
        "查下",
        "搜一下",
        "搜索",
        "帮我查",
        "请问",
        "最近",
        "近期",
        "最新",
        "新闻",
        "消息",
        "今天",
        "今日",
        "现在",
        "目前",
        "在不在",
        "在哪",
        "是否",
        "一下",
    )
    cjk_stopwords = {"新闻", "消息", "今天", "今日", "最近", "近期", "一下", "查下", "搜索"}

    seen: set[str] = set()
    tokens: list[str] = []

    def add(token: str) -> None:
        normalized = str(token or "").strip().lower()
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        tokens.append(normalized)

    for part in re.split(r"[^a-z0-9]+", text.lower()):
        if len(part) < 2 or part in ascii_stopwords:
            continue
        add(part)

    cjk_text = text
    for filler in cjk_fillers:
        cjk_text = cjk_text.replace(filler, " ")

    for segment in re.findall(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]+", cjk_text):
        cleaned = segment.strip()
        if len(cleaned) < 2:
            continue
        max_size = min(4, len(cleaned))
        for size in range(max_size, 1, -1):
            for start in range(0, len(cleaned) - size + 1):
                token = cleaned[start : start + size]
                if token in cjk_stopwords:
                    continue
                add(token)
                if len(tokens) >= 24:
                    return tokens

    return tokens


def _query_relevance_score(query: str, item: dict[str, Any]) -> float:
    title = str(item.get("title") or "").lower()
    snippet = str(item.get("snippet") or "").lower()
    domain = str(item.get("domain") or "").lower()
    tokens = _tokenize_query(query)
    score = 0.0
    for token in tokens:
        weight = 1.0 + min(len(token), 4) * 0.35
        if token in title:
            score += 4.0 * weight
        if token in snippet:
            score += 2.0 * weight
        if token.isascii() and token in domain:
            score += 1.5 * weight
    return score


def _score_web_result(query: str, item: dict[str, Any]) -> float:
    domain = str(item.get("domain") or "").lower()
    score = _query_relevance_score(query, item)
    if domain.endswith(".gov") or domain.endswith(".edu"):
        score += 2.5
    if any(flag in domain for flag in ("official", "docs", "developer", "openai.com", "github.com")):
        score += 1.5
    if item.get("published_at"):
        score += 0.5
    return score


def _query_looks_specific(query: str) -> bool:
    text = (query or "").strip()
    if not text:
        return False

    normalized = text.lower()
    generic_markers = (
        "news",
        "latest",
        "recent",
        "today",
        "score",
        "scores",
        "baseball",
        "mlb",
        "npb",
        "kbo",
        "棒球",
        "野球",
        "新闻",
        "消息",
        "最近",
        "近期",
        "今天",
        "今日",
        "查一下",
        "查下",
        "搜一下",
        "搜索",
        "在不在",
        "是否",
    )
    for marker in generic_markers:
        normalized = normalized.replace(marker, " ")

    ascii_tokens = [part for part in re.split(r"[^a-z0-9]+", normalized) if len(part) >= 2]
    cjk_chars = "".join(re.findall(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]+", normalized))
    return bool(ascii_tokens or len(cjk_chars) >= 2)


def _looks_like_script_payload(text: str) -> bool:
    sample = (text or "")[:6000].lower()
    if not sample:
        return False

    if "sourcemappingurl=" in sample:
        return True

    markers = [
        "function(",
        "var ",
        "const ",
        "let ",
        "window.",
        "document.",
        "=>",
    ]
    hits = sum(1 for m in markers if m in sample)
    longest_line = max((len(line) for line in sample.splitlines()), default=0)
    punct = sum(ch in "{}[]();=<>/\\*" for ch in sample)
    alpha = sum(ch.isalpha() for ch in sample) or 1
    punct_ratio = punct / alpha

    return (hits >= 3 and longest_line >= 220) or punct_ratio >= 0.45


def _extract_search_query(url: str) -> str | None:
    try:
        parsed = urllib.parse.urlsplit(url)
    except Exception:
        return None

    host = (parsed.hostname or "").lower()
    if not host:
        return None

    q = urllib.parse.parse_qs(parsed.query or "")
    key = None
    if "google." in host or "bing." in host:
        key = "q"
    elif "yahoo." in host:
        key = "p"
    elif "baidu." in host:
        key = "wd"

    if not key:
        return None
    vals = q.get(key) or []
    if not vals:
        return None
    out = (vals[0] or "").strip()
    return out or None


def _clean_html_fragment(raw_html: str) -> str:
    text = re.sub(r"(?s)<[^>]+>", " ", raw_html or "")
    text = unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _decode_ddg_redirect(raw_url: str) -> str:
    if not raw_url:
        return raw_url
    url = unescape(raw_url).strip()
    absolute = urllib.parse.urljoin("https://duckduckgo.com", url)
    try:
        parsed = urllib.parse.urlsplit(absolute)
    except Exception:
        return absolute

    host = (parsed.hostname or "").lower()
    if host.endswith("duckduckgo.com") and parsed.path == "/l/":
        q = urllib.parse.parse_qs(parsed.query or "")
        target = (q.get("uddg") or [""])[0].strip()
        if target:
            return urllib.parse.unquote(target)
    return absolute


def _extract_ddg_results(raw_html: str, max_results: int) -> list[dict[str, str]]:
    html = raw_html or ""
    limit = max(1, min(20, int(max_results)))
    patterns = [
        re.compile(
            r'(?is)<a[^>]*class="[^"]*result__a[^"]*"[^>]*href="([^"]+)"[^>]*>(.*?)</a>'
        ),
        re.compile(
            r"(?is)<a[^>]*class=['\"][^'\"]*result-link[^'\"]*['\"][^>]*href=['\"]([^'\"]+)['\"][^>]*>(.*?)</a>"
        ),
    ]

    seen: set[str] = set()
    out: list[dict[str, str]] = []

    for pattern in patterns:
        for match in pattern.finditer(html):
            href = _decode_ddg_redirect(match.group(1) or "")
            title = _clean_html_fragment(match.group(2) or "")
            if not href or not title:
                continue
            try:
                parsed = urllib.parse.urlsplit(href)
            except Exception:
                continue
            if parsed.scheme not in {"http", "https"}:
                continue
            host = (parsed.hostname or "").lower()
            if host.endswith("duckduckgo.com") and parsed.path == "/y.js":
                continue

            key = f"{href}|{title}".lower()
            if key in seen:
                continue
            seen.add(key)

            snippet = ""
            window = html[match.end() : match.end() + 2400]
            snippet_match = re.search(
                r'(?is)<(?:a|div|span)[^>]*class="[^"]*result__snippet[^"]*"[^>]*>(.*?)</(?:a|div|span)>',
                window,
            )
            if not snippet_match:
                snippet_match = re.search(
                    r"(?is)<td[^>]*class=['\"][^'\"]*result-snippet[^'\"]*['\"][^>]*>(.*?)</td>",
                    window,
                )
            if snippet_match:
                snippet = _clean_html_fragment(snippet_match.group(1) or "")

            out.append({"title": title, "url": href, "snippet": snippet})
            if len(out) >= limit:
                return out

    return out


def _looks_news_like_query(query: str) -> bool:
    text = (query or "").strip().lower()
    if not text:
        return False
    keywords = [
        "news",
        "latest",
        "recent",
        "breaking",
        "headline",
        "headlines",
        "today",
        "score",
        "scores",
        "最近",
        "近期",
        "近况",
        "新闻",
        "消息",
        "今日",
        "今天",
        "速报",
        "戰報",
        "战报",
        "比分",
        "ニュース",
    ]
    return any(k in text for k in keywords)


def _looks_baseball_query(query: str) -> bool:
    text = (query or "").strip().lower()
    if not text:
        return False
    keywords = [
        "baseball",
        "mlb",
        "npb",
        "kbo",
        "棒球",
        "野球",
        "甲子園",
        "甲子园",
        "大谷",
    ]
    return any(k in text for k in keywords)


def _build_rss_candidates(query: str) -> list[tuple[str, str]]:
    q = (query or "").strip()
    out: list[tuple[str, str]] = []
    is_baseball = _looks_baseball_query(q)
    query_has_cjk = bool(re.search(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]", q))
    query_is_specific = _query_looks_specific(q)

    if is_baseball:
        q_en = urllib.parse.quote_plus(f"{q} baseball")
        q_ja = urllib.parse.quote_plus(f"{q} 野球")
        google_news = [
            (
                "google_news_baseball_ja",
                f"https://news.google.com/rss/search?q={q_ja}&hl=ja&gl=JP&ceid=JP:ja",
            ),
            (
                "google_news_baseball_en",
                f"https://news.google.com/rss/search?q={q_en}&hl=en-US&gl=US&ceid=US:en",
            ),
        ]
        if not query_has_cjk:
            google_news.reverse()

        generic_feeds = [
            ("mlb_official_rss", "https://www.mlb.com/feeds/news/rss.xml"),
            ("espn_mlb_rss", "https://www.espn.com/espn/rss/mlb/news"),
            ("yahoo_mlb_rss", "https://sports.yahoo.com/mlb/rss/"),
            ("nhk_sports_rss", "https://www3.nhk.or.jp/rss/news/cat7.xml"),
        ]
        if query_is_specific:
            out.extend(google_news)
            out.extend(generic_feeds)
        else:
            out.extend(generic_feeds)
            out.extend(google_news)
    else:
        quoted = urllib.parse.quote_plus(q)
        out.append(
            (
                "google_news_query_zh",
                f"https://news.google.com/rss/search?q={quoted}&hl=zh-CN&gl=CN&ceid=CN:zh-Hans",
            )
        )
        out.append(
            (
                "google_news_query_en",
                f"https://news.google.com/rss/search?q={quoted}&hl=en-US&gl=US&ceid=US:en",
            )
        )

    seen: set[str] = set()
    deduped: list[tuple[str, str]] = []
    for name, url in out:
        if url in seen:
            continue
        seen.add(url)
        deduped.append((name, url))
    return deduped


def _extract_google_news_rss_results(raw_xml: str, max_results: int) -> list[dict[str, str]]:
    limit = max(1, min(20, int(max_results)))
    xml_text = (raw_xml or "").strip()
    if not xml_text:
        return []

    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []

    items = root.findall(".//item")
    out: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in items:
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        desc = (item.findtext("description") or "").strip()
        if not title or not link:
            continue

        title = _clean_html_fragment(title)
        link = _clean_html_fragment(link)
        snippet = _clean_html_fragment(desc)
        key = f"{title}|{link}".lower()
        if key in seen:
            continue
        seen.add(key)
        published_at = (item.findtext("pubDate") or "").strip()
        entry = {"title": title, "url": link, "snippet": snippet}
        if published_at:
            entry["published_at"] = published_at
        out.append(entry)
        if len(out) >= limit:
            break
    return out


def _normalize_url_for_request(raw_url: str) -> str:
    """
    Make URL safe for urllib by encoding non-ASCII host/path/query.
    """
    url = (raw_url or "").strip()
    parsed = urllib.parse.urlsplit(url)

    scheme = (parsed.scheme or "").lower()
    if scheme not in {"http", "https"}:
        raise ValueError("Only http/https URLs are supported")
    if not parsed.netloc:
        raise ValueError("Invalid URL")

    host = parsed.hostname or ""
    if not host:
        raise ValueError("Invalid URL")
    host_ascii = host.encode("idna").decode("ascii")

    auth = ""
    if parsed.username is not None:
        auth = urllib.parse.quote(parsed.username, safe="")
        if parsed.password is not None:
            auth += ":" + urllib.parse.quote(parsed.password, safe="")
        auth += "@"

    port = f":{parsed.port}" if parsed.port else ""
    netloc = f"{auth}{host_ascii}{port}"

    path = urllib.parse.quote(urllib.parse.unquote(parsed.path or "/"), safe="/%:@!$&'()*+,;=-._~")
    query = urllib.parse.quote(urllib.parse.unquote(parsed.query or ""), safe="=&%:@!$'()*+,;/-._~")

    return urllib.parse.urlunsplit((scheme, netloc, path, query, ""))


def _is_cert_verify_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "certificate_verify_failed" in text or "certificate verify failed" in text


def _normalize_search_query(query: str) -> str:
    return re.sub(r"\s+", " ", (query or "").strip())


def _expand_search_variants(query: str) -> list[str]:
    normalized = _normalize_search_query(query)
    if not normalized:
        return []

    variants: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        clean = _normalize_search_query(value)
        key = clean.lower()
        if not clean or key in seen:
            return
        seen.add(key)
        variants.append(clean)

    add(normalized)

    hex_patterns = list(re.finditer(r"(?i)\b(?:0x([0-9a-f]{1,4})|([0-9a-f]{1,4})h)\b", normalized))
    for match in hex_patterns:
        digits = (match.group(1) or match.group(2) or "").upper()
        if not digits:
            continue
        token_variants = [f"{digits}h", f"{digits} h", f"0x{digits}"]
        for token in token_variants:
            add(token)
            replaced = normalized[: match.start()] + token + normalized[match.end() :]
            add(replaced)

    return variants


def _build_search_pattern(query: str) -> re.Pattern[str] | None:
    normalized = _normalize_search_query(query)
    if not normalized:
        return None
    parts = [re.escape(part) for part in normalized.split(" ") if part]
    if not parts:
        return None
    body = r"\s+".join(parts)
    if len(parts) == 1 and re.fullmatch(r"(?i)(?:0x)?[0-9a-f]{1,4}h?", normalized):
        body = rf"(?<![0-9A-F]){body}(?![0-9A-F])"
    return re.compile(body, flags=re.IGNORECASE)


def _page_hint_for_offset(text: str, offset: int) -> int | None:
    page = None
    for match in re.finditer(r"--- Page (\d+) ---", text):
        if match.start() > offset:
            break
        try:
            page = int(match.group(1))
        except Exception:
            page = None
    return page


def _spans_overlap(left: tuple[int, int], right: tuple[int, int]) -> bool:
    return left[0] < right[1] and right[0] < left[1]


def _looks_like_pdf_path(path: Path) -> bool:
    if path.suffix.lower() == ".pdf":
        return True
    try:
        with path.open("rb") as fp:
            return fp.read(5).startswith(b"%PDF-")
    except Exception:
        return False


def _xlsx_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        try:
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
        except Exception:
            pass
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return str(value.isoformat())
        except Exception:
            pass
    return str(value).strip()


def _heading_score(query_norm: str, heading_norm: str) -> float:
    if not query_norm or not heading_norm:
        return 0.0
    if query_norm == heading_norm:
        return 10.0
    if query_norm in heading_norm:
        return 7.0 + len(query_norm) / max(1, len(heading_norm))
    if heading_norm in query_norm:
        return 6.0 + len(heading_norm) / max(1, len(query_norm))
    query_tokens = set(query_norm.split())
    heading_tokens = set(heading_norm.split())
    if not query_tokens or not heading_tokens:
        return 0.0
    overlap = len(query_tokens & heading_tokens)
    if not overlap:
        return 0.0
    return overlap / max(1, len(query_tokens | heading_tokens))


def _find_best_heading(
    headings: list[dict[str, object]],
    query: str,
) -> dict[str, object] | None:
    query_norm = normalize_lookup_text(query)
    best: tuple[float, dict[str, object] | None] = (0.0, None)
    for heading in headings:
        heading_norm = str(heading.get("normalized") or "")
        score = _heading_score(query_norm, heading_norm)
        if score > best[0]:
            best = (score, heading)
    if best[0] <= 0.0:
        return None
    return best[1]


def _line_matches_heading(line: str, heading: dict[str, object]) -> bool:
    line_norm = normalize_lookup_text(line)
    heading_norm = str(heading.get("normalized") or "")
    return bool(line_norm and heading_norm and _heading_score(line_norm, heading_norm) >= 6.0)


def _extract_section_from_pdf_pages(
    pages: list[tuple[int, str]],
    headings: list[dict[str, object]],
    heading_query: str,
    max_chars: int,
) -> dict[str, Any]:
    match = _find_best_heading(headings, heading_query)
    if not match:
        return {"ok": False, "error": f"Heading not found: {heading_query}"}

    ordered = sorted(headings, key=lambda item: (int(item.get("page") or 0), int(item.get("line_index") or 0)))
    match_idx = ordered.index(match)
    next_heading = ordered[match_idx + 1] if match_idx + 1 < len(ordered) else None

    collecting = False
    chunks: list[str] = []
    total = 0
    page_start = int(match.get("page") or 0)
    page_end = page_start

    for page_num, body in pages:
        if page_num < page_start:
            continue
        lines = body.splitlines()
        started_here = False
        for line_idx, line in enumerate(lines, start=1):
            if not collecting:
                if page_num == page_start and _line_matches_heading(line, match):
                    collecting = True
                    started_here = True
                else:
                    continue
            if (
                next_heading
                and page_num == int(next_heading.get("page") or 0)
                and _line_matches_heading(line, next_heading)
                and not (started_here and page_num == page_start and line_idx == int(match.get("line_index") or 0))
            ):
                collecting = False
                break
            line_text = line.rstrip()
            if not line_text:
                continue
            chunks.append(line_text)
            total += len(line_text) + 1
            page_end = page_num
            if total >= max_chars:
                collecting = False
                break
        if not collecting and chunks:
            break

    content = truncate_text("\n".join(chunks).strip(), max(512, int(max_chars)))
    return {
        "ok": True,
        "matched_heading": str(match.get("heading") or heading_query),
        "matched_section": str(match.get("section") or ""),
        "page_start": page_start,
        "page_end": page_end,
        "content": content,
    }


def _derive_fact_check_queries(claim: str) -> list[str]:
    text = (claim or "").strip()
    if not text:
        return []
    queries: list[str] = []
    seen: set[str] = set()

    def add(value: str) -> None:
        normalized = _normalize_search_query(value)
        key = normalized.lower()
        if not normalized or key in seen:
            return
        seen.add(key)
        queries.append(normalized)

    for match in re.finditer(r"(?i)\b(?:0x[0-9a-f]{1,4}|[0-9a-f]{1,4}h)\b", text):
        add(match.group(0))
    for match in re.finditer(r'"([^"]+)"|“([^”]+)”|\'([^\']+)\'', text):
        for group in match.groups():
            if group:
                add(group)
    for match in re.finditer(r"\b\d+(?:\.\d+){1,5}\b", text):
        add(match.group(0))

    tokens = [
        token
        for token in re.findall(r"[A-Za-z][A-Za-z0-9/_-]{3,}", text)
        if token.lower() not in {"that", "with", "from", "there", "which", "this", "does", "have"}
    ]
    for token in tokens[:4]:
        add(token)
    if not queries:
        add(text)
    return queries[:8]


def _is_negative_claim(claim: str) -> bool:
    text = (claim or "").strip().lower()
    markers = (
        " not ",
        " no ",
        "none",
        "without",
        "does not",
        "is not",
        "isn't",
        "没有",
        "不存在",
        "未找到",
        "不是",
        "不支持",
    )
    padded = f" {text} "
    return any(marker in padded for marker in markers)


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w.\-]+", "_", (name or "").strip(), flags=re.UNICODE).strip("._")
    if not cleaned:
        cleaned = "download.bin"
    return cleaned[:180]


def _guess_filename_from_response(url: str, content_type: str, content_disposition: str) -> str:
    cd = content_disposition or ""
    filename_match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, flags=re.IGNORECASE)
    if filename_match:
        name = urllib.parse.unquote(filename_match.group(1) or "").strip()
        if name:
            return _safe_filename(name)

    parsed = urllib.parse.urlsplit(url)
    candidate = Path(urllib.parse.unquote(parsed.path or "")).name
    if candidate:
        return _safe_filename(candidate)

    ct = (content_type or "").lower()
    if "application/pdf" in ct:
        return "download.pdf"
    if "application/json" in ct:
        return "download.json"
    if "text/html" in ct:
        return "download.html"
    if "text/plain" in ct:
        return "download.txt"
    return "download.bin"


class LocalToolExecutor:
    def __init__(self, config: AppConfig) -> None:
        self.config = config
        self._runtime_ctx = threading.local()
        self._web_cache_lock = threading.Lock()
        self._web_cache_dir = (config.workspace_root / "app" / "data" / "web_cache").resolve()
        self._web_cache_dir.mkdir(parents=True, exist_ok=True)
        self._docker_sandbox = DockerSandboxManager(
            workspace_root=config.workspace_root,
            allowed_roots=config.allowed_roots,
            docker_bin=config.docker_bin,
            image=config.docker_image,
            network=config.docker_network,
            memory=config.docker_memory,
            cpus=config.docker_cpus,
            pids_limit=config.docker_pids_limit,
            container_prefix=config.docker_container_prefix,
        )

    def set_runtime_context(self, *, execution_mode: str | None = None, session_id: str | None = None) -> None:
        mode = (execution_mode or "").strip().lower()
        if mode not in {"host", "docker"}:
            mode = self.config.execution_mode
        self._runtime_ctx.execution_mode = mode
        sid = str(session_id or "").strip() or "__anon__"
        self._runtime_ctx.session_id = sid

    def clear_runtime_context(self) -> None:
        for key in ("execution_mode", "session_id"):
            try:
                delattr(self._runtime_ctx, key)
            except Exception:
                pass

    def _current_execution_mode(self) -> str:
        mode = str(getattr(self._runtime_ctx, "execution_mode", "") or "").strip().lower()
        if mode in {"host", "docker"}:
            return mode
        return self.config.execution_mode

    def _current_session_id(self) -> str:
        return str(getattr(self._runtime_ctx, "session_id", "") or "__anon__")

    def _web_cache_path(self, prefix: str, payload: dict[str, Any]) -> Path:
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        digest = hashlib.sha256(raw).hexdigest()
        return self._web_cache_dir / f"{prefix}_{digest}.json"

    def _load_web_cache(self, prefix: str, payload: dict[str, Any], max_age_sec: int = 900) -> dict[str, Any] | None:
        path = self._web_cache_path(prefix, payload)
        if not path.is_file():
            return None
        try:
            with self._web_cache_lock:
                cached = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return None
        saved_at = float(cached.get("saved_at") or 0)
        if saved_at <= 0 or (time.time() - saved_at) > max_age_sec:
            return None
        payload_data = cached.get("payload")
        return payload_data if isinstance(payload_data, dict) else None

    def _save_web_cache(self, prefix: str, payload: dict[str, Any], result: dict[str, Any]) -> None:
        path = self._web_cache_path(prefix, payload)
        body = {
            "saved_at": time.time(),
            "payload": result,
        }
        try:
            with self._web_cache_lock:
                path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
        except Exception:
            return

    def docker_available(self) -> bool:
        return self._docker_sandbox.docker_available()

    def docker_status(self) -> tuple[bool, str]:
        ok = self._docker_sandbox.docker_available()
        return ok, self._docker_sandbox.docker_status_message()

    @property
    def tool_specs(self) -> list[dict[str, Any]]:
        return [
            {
                "type": "function",
                "name": "run_shell",
                "description": "Run a safe shell command in workspace. Supports simple commands without pipes.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "command": {"type": "string", "description": "Shell command, e.g. `ls -la` or `rg TODO .`"},
                        "cwd": {"type": "string", "description": "Working directory relative to workspace", "default": "."},
                        "timeout_sec": {"type": "integer", "minimum": 1, "maximum": 120, "default": 15},
                    },
                    "required": ["command"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "list_directory",
                "description": "List files in a workspace directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "default": "."},
                        "max_entries": {"type": "integer", "minimum": 1, "maximum": 500, "default": 200},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "read_text_file",
                "description": (
                    "Read a local text/document file in allowed roots. "
                    "For PDF/DOCX/MSG/XLSX it auto-extracts text; supports chunked reads with start_char."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "start_char": {"type": "integer", "minimum": 0, "default": 0},
                        "max_chars": {"type": "integer", "minimum": 128, "maximum": 1000000, "default": 200000},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "search_text_in_file",
                "description": (
                    "Search within a local text/document file and return matching evidence snippets. "
                    "For PDF/DOCX/MSG/XLSX it searches extracted text and automatically expands hex variants like 15h/15 h/0x15."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "query": {"type": "string"},
                        "max_matches": {"type": "integer", "minimum": 1, "maximum": 20, "default": 8},
                        "context_chars": {"type": "integer", "minimum": 40, "maximum": 2000, "default": 280},
                    },
                    "required": ["path", "query"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "multi_query_search",
                "description": "Run multiple file-search queries against one local file and merge the evidence snippets.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "queries": {"type": "array", "items": {"type": "string"}},
                        "per_query_max_matches": {"type": "integer", "minimum": 1, "maximum": 10, "default": 3},
                        "context_chars": {"type": "integer", "minimum": 40, "maximum": 2000, "default": 280},
                    },
                    "required": ["path", "queries"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "doc_index_build",
                "description": "Build or inspect a cached document index for a local PDF, including headings and cache status.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "force_rebuild": {"type": "boolean", "default": False},
                        "max_headings": {"type": "integer", "minimum": 20, "maximum": 2000, "default": 400},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "read_section_by_heading",
                "description": "Read a document section by matching a heading or section number and returning that section's content.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "heading": {"type": "string"},
                        "max_chars": {"type": "integer", "minimum": 512, "maximum": 50000, "default": 12000},
                    },
                    "required": ["path", "heading"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "table_extract",
                "description": "Extract tables from a local PDF or XLSX file, optionally narrowed by query or page hint.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "query": {"type": "string", "default": ""},
                        "page_hint": {"type": "integer", "minimum": 1, "default": 0},
                        "max_tables": {"type": "integer", "minimum": 1, "maximum": 20, "default": 5},
                        "max_rows": {"type": "integer", "minimum": 1, "maximum": 200, "default": 25},
                    },
                    "required": ["path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "fact_check_file",
                "description": "Check whether a file provides supporting evidence for a claim, using derived or provided search queries.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "claim": {"type": "string"},
                        "queries": {"type": "array", "items": {"type": "string"}, "default": []},
                        "max_evidence": {"type": "integer", "minimum": 1, "maximum": 12, "default": 6},
                    },
                    "required": ["path", "claim"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "search_codebase",
                "description": "Search code or text files under a local root and return file, line, and context matches.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "root": {"type": "string", "default": "."},
                        "max_matches": {"type": "integer", "minimum": 1, "maximum": 100, "default": 20},
                        "file_glob": {"type": "string", "default": ""},
                        "use_regex": {"type": "boolean", "default": False},
                        "case_sensitive": {"type": "boolean", "default": False},
                    },
                    "required": ["query"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "copy_file",
                "description": "Copy a file (binary-safe) from src_path to dst_path in allowed roots.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "src_path": {"type": "string"},
                        "dst_path": {"type": "string"},
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                    },
                    "required": ["src_path", "dst_path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "extract_zip",
                "description": (
                    "Extract a local .zip archive into a target directory in allowed roots. "
                    "Supports overwrite controls and safety limits."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "zip_path": {"type": "string", "description": "Path to local .zip file"},
                        "dst_dir": {
                            "type": "string",
                            "description": "Destination directory. Empty means same name next to zip file.",
                            "default": "",
                        },
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                        "max_entries": {"type": "integer", "minimum": 1, "maximum": 100000, "default": 20000},
                        "max_total_bytes": {
                            "type": "integer",
                            "minimum": 1024,
                            "maximum": 2147483648,
                            "default": 524288000,
                        },
                    },
                    "required": ["zip_path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "extract_msg_attachments",
                "description": (
                    "Extract attachments from a local Outlook .msg email into a target directory in allowed roots. "
                    "Use this to open/read Excel/image attachments referenced inside .msg."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "msg_path": {"type": "string", "description": "Path to local .msg file"},
                        "dst_dir": {
                            "type": "string",
                            "description": "Destination directory. Empty means <msg_stem>_attachments next to msg file.",
                            "default": "",
                        },
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                        "max_attachments": {"type": "integer", "minimum": 1, "maximum": 5000, "default": 500},
                        "max_total_bytes": {
                            "type": "integer",
                            "minimum": 1024,
                            "maximum": 2147483648,
                            "default": 524288000,
                        },
                    },
                    "required": ["msg_path"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "write_text_file",
                "description": "Create or overwrite a UTF-8 text file in workspace.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "content": {"type": "string"},
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                    },
                    "required": ["path", "content"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "append_text_file",
                "description": "Append UTF-8 text to an existing file (or create new file) in workspace.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "content": {"type": "string"},
                        "create_if_missing": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                    },
                    "required": ["path", "content"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "replace_in_file",
                "description": "Replace target text in a UTF-8 text file in workspace.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "old_text": {"type": "string"},
                        "new_text": {"type": "string"},
                        "replace_all": {"type": "boolean", "default": False},
                        "max_replacements": {"type": "integer", "minimum": 1, "maximum": 200, "default": 1},
                    },
                    "required": ["path", "old_text", "new_text"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "fetch_web",
                "description": "Fetch web content from a URL for information lookup.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string", "description": "http/https URL"},
                        "max_chars": {"type": "integer", "minimum": 512, "maximum": 500000, "default": 120000},
                        "timeout_sec": {"type": "integer", "minimum": 3, "maximum": 30, "default": 12},
                    },
                    "required": ["url"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "download_web_file",
                "description": (
                    "Download a web file (binary-safe) and save it to local path under allowed roots. "
                    "Use this when user asks to download/save PDF/ZIP/images/binaries."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string", "description": "http/https URL"},
                        "dst_path": {
                            "type": "string",
                            "description": (
                                "Destination file path. If empty, auto-saves to downloads/<filename> under workspace."
                            ),
                            "default": "",
                        },
                        "overwrite": {"type": "boolean", "default": True},
                        "create_dirs": {"type": "boolean", "default": True},
                        "timeout_sec": {"type": "integer", "minimum": 3, "maximum": 120, "default": 20},
                        "max_bytes": {"type": "integer", "minimum": 1024, "maximum": 209715200, "default": 52428800},
                    },
                    "required": ["url"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "search_web",
                "description": "Search the web by query and return candidate URLs/snippets.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "Search query keywords"},
                        "max_results": {"type": "integer", "minimum": 1, "maximum": 20, "default": 5},
                        "timeout_sec": {"type": "integer", "minimum": 3, "maximum": 30, "default": 12},
                    },
                    "required": ["query"],
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "list_sessions",
                "description": "List recent local chat sessions so the agent can locate past context.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "max_sessions": {"type": "integer", "minimum": 1, "maximum": 200, "default": 20},
                    },
                    "additionalProperties": False,
                },
            },
            {
                "type": "function",
                "name": "read_session_history",
                "description": "Read one local chat session's summary and recent turns by session_id.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "session_id": {"type": "string"},
                        "max_turns": {"type": "integer", "minimum": 1, "maximum": 800, "default": 80},
                    },
                    "required": ["session_id"],
                    "additionalProperties": False,
                },
            },
        ]

    def execute(self, name: str, arguments: dict[str, Any]) -> dict[str, Any]:
        if name == "run_shell":
            return self.run_shell(**arguments)
        if name == "list_directory":
            return self.list_directory(**arguments)
        if name == "read_text_file":
            return self.read_text_file(**arguments)
        if name == "search_text_in_file":
            return self.search_text_in_file(**arguments)
        if name == "multi_query_search":
            return self.multi_query_search(**arguments)
        if name == "doc_index_build":
            return self.doc_index_build(**arguments)
        if name == "read_section_by_heading":
            return self.read_section_by_heading(**arguments)
        if name == "table_extract":
            return self.table_extract(**arguments)
        if name == "fact_check_file":
            return self.fact_check_file(**arguments)
        if name == "search_codebase":
            return self.search_codebase(**arguments)
        if name == "copy_file":
            return self.copy_file(**arguments)
        if name == "extract_zip":
            return self.extract_zip(**arguments)
        if name == "extract_msg_attachments":
            return self.extract_msg_attachments(**arguments)
        if name == "write_text_file":
            return self.write_text_file(**arguments)
        if name == "append_text_file":
            return self.append_text_file(**arguments)
        if name == "replace_in_file":
            return self.replace_in_file(**arguments)
        if name == "fetch_web":
            return self.fetch_web(**arguments)
        if name == "download_web_file":
            return self.download_web_file(**arguments)
        if name == "search_web":
            return self.search_web(**arguments)
        if name == "list_sessions":
            return self.list_sessions(**arguments)
        if name == "read_session_history":
            return self.read_session_history(**arguments)
        return {"ok": False, "error": f"Unknown tool: {name}"}

    def run_shell(self, command: str, cwd: str = ".", timeout_sec: int = 15) -> dict[str, Any]:
        try:
            argv = shlex.split(command)
        except Exception as exc:
            return {"ok": False, "error": f"Command parse failed: {exc}"}

        if not argv:
            return {"ok": False, "error": "Empty command"}

        if any(token in command for token in ["|", "&&", "||", ";", "$(", "`"]):
            return {
                "ok": False,
                "error": "Complex shell operators are blocked for safety. Use a single command only.",
            }

        execution_mode = self._current_execution_mode()
        session_id = self._current_session_id()
        timeout_val = max(1, min(120, timeout_sec))

        # Docker image commonly exposes python3, while users may type python.
        if execution_mode == "docker" and argv[0] == "python":
            argv[0] = "python3"

        base_cmd = argv[0]
        if base_cmd not in self.config.allowed_commands:
            return {
                "ok": False,
                "error": f"Command not allowed: {base_cmd}. Allowed: {', '.join(self.config.allowed_commands)}",
            }

        try:
            real_cwd = _resolve_workspace_path(self.config, cwd)
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

        timeout_val = max(1, min(120, timeout_sec))
        execution_mode = self._current_execution_mode()
        session_id = self._current_session_id()

        try:
            sandbox_cwd = None
            mounts: list[dict[str, str]] = []
            if execution_mode == "docker":
                try:
                    sandbox_cwd = self._docker_sandbox.container_path_for_host(real_cwd)
                except Exception:
                    sandbox_cwd = None
                mounts = self._docker_sandbox.mount_mappings()
                proc = self._docker_sandbox.run_in_sandbox(
                    session_id=session_id,
                    argv=argv,
                    cwd=real_cwd,
                    timeout_sec=timeout_val,
                    container_cwd=sandbox_cwd,
                )
            else:
                proc = subprocess.run(
                    argv,
                    cwd=str(real_cwd),
                    capture_output=True,
                    text=True,
                    timeout=timeout_val,
                    check=False,
                )
            payload: dict[str, Any] = {
                "ok": proc.returncode == 0,
                "returncode": proc.returncode,
                "stdout": _truncate_output(proc.stdout),
                "stderr": _truncate_output(proc.stderr),
                "cwd": str(real_cwd),
                "host_cwd": str(real_cwd),
                "command": " ".join(shlex.quote(x) for x in argv),
                "execution_mode": execution_mode,
            }
            if execution_mode == "docker":
                payload["sandbox_cwd"] = sandbox_cwd or ""
                payload["mount_mappings"] = mounts
                payload["path_mapping_hint"] = (
                    "Files in sandbox_cwd are persisted to host_cwd via Docker bind mounts. "
                    "Always report host_cwd/host absolute paths to user."
                )
            return payload
        except subprocess.TimeoutExpired:
            return {"ok": False, "error": f"Command timed out after {timeout_val}s"}
        except Exception as exc:
            return {"ok": False, "error": f"run_shell failed: {exc}"}

    def list_directory(self, path: str = ".", max_entries: int = 200) -> dict[str, Any]:
        try:
            real_path = _resolve_workspace_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_dir():
                return {"ok": False, "error": f"Not a directory: {path}"}

            entries = []
            for child in sorted(real_path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
                entries.append(
                    {
                        "name": child.name,
                        "is_dir": child.is_dir(),
                        "size": child.stat().st_size if child.is_file() else None,
                    }
                )
                if len(entries) >= max_entries:
                    break
            return {"ok": True, "path": str(real_path), "entries": entries}
        except Exception as exc:
            return {"ok": False, "error": f"list_directory failed: {exc}"}

    def list_sessions(self, max_sessions: int = 20) -> dict[str, Any]:
        try:
            limit = max(1, min(200, int(max_sessions)))
            rows: list[dict[str, Any]] = []
            files = sorted(
                self.config.sessions_dir.glob("*.json"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            for path in files[:limit]:
                try:
                    payload = json.loads(path.read_text(encoding="utf-8"))
                except Exception:
                    continue
                sid = str(payload.get("id") or path.stem)
                turns = payload.get("turns", [])
                if not isinstance(turns, list):
                    turns = []
                title = "新会话"
                preview = ""
                for turn in turns:
                    if not isinstance(turn, dict):
                        continue
                    role = str(turn.get("role") or "")
                    text = str(turn.get("text") or "").strip()
                    if role == "user" and text:
                        title = text.replace("\n", " ")[:60]
                        break
                if turns:
                    last = turns[-1]
                    if isinstance(last, dict):
                        preview = str(last.get("text") or "").replace("\n", " ").strip()[:120]
                rows.append(
                    {
                        "session_id": sid,
                        "title": title,
                        "preview": preview,
                        "turn_count": len(turns),
                        "updated_at": str(payload.get("updated_at") or ""),
                        "created_at": str(payload.get("created_at") or ""),
                    }
                )
            return {"ok": True, "count": len(rows), "sessions": rows}
        except Exception as exc:
            return {"ok": False, "error": f"list_sessions failed: {exc}"}

    def read_session_history(self, session_id: str, max_turns: int = 80) -> dict[str, Any]:
        sid = str(session_id or "").strip()
        if not sid:
            return {"ok": False, "error": "session_id cannot be empty"}
        if "/" in sid or "\\" in sid or ".." in sid:
            return {"ok": False, "error": "Invalid session_id"}
        try:
            session_path = (self.config.sessions_dir / f"{sid}.json").resolve()
            if not _is_within(session_path, self.config.sessions_dir):
                return {"ok": False, "error": "Invalid session path"}
            if not session_path.exists():
                return {"ok": False, "error": f"Session not found: {sid}"}
            payload = json.loads(session_path.read_text(encoding="utf-8"))
            turns = payload.get("turns", [])
            if not isinstance(turns, list):
                turns = []
            keep = max(1, min(800, int(max_turns)))
            sliced = turns[-keep:]
            trimmed_turns: list[dict[str, str]] = []
            for turn in sliced:
                if not isinstance(turn, dict):
                    continue
                trimmed_turns.append(
                    {
                        "role": str(turn.get("role") or "user"),
                        "text": str(turn.get("text") or ""),
                        "created_at": str(turn.get("created_at") or ""),
                    }
                )
            return {
                "ok": True,
                "session_id": sid,
                "summary": str(payload.get("summary") or ""),
                "turn_count": len(turns),
                "turns": trimmed_turns,
            }
        except Exception as exc:
            return {"ok": False, "error": f"read_session_history failed: {exc}"}

    def read_text_file(self, path: str, start_char: int = 0, max_chars: int = 200000) -> dict[str, Any]:
        try:
            real_path = _resolve_source_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}
            suffix = real_path.suffix.lower()
            source_format = "text_utf8"
            full_text = ""

            # For office/binary documents, try structured extraction first
            # so users can "download then read" in one flow.
            if suffix == ".pdf":
                source_format = "pdf_text_extracted"
                try:
                    full_text = extract_pdf_text_from_path(real_path, max_chars=1_000_000)
                except Exception as exc:
                    full_text = f"[文档解析失败: {exc}]"
            elif suffix in {
                ".docx",
                ".msg",
                ".xlsx",
                ".xlsm",
                ".xltx",
                ".xltm",
                ".xls",
                ".pptx",
                ".pptm",
                ".ppt",
                ".atom",
                ".rss",
                ".xml",
            }:
                from app.attachments import extract_document_text  # lazy import

                extracted = extract_document_text(str(real_path), max_chars=1_000_000) or ""
                full_text = extracted
                if suffix == ".docx":
                    source_format = "docx_text_extracted"
                elif suffix == ".msg":
                    source_format = "msg_text_extracted"
                elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
                    source_format = "xlsx_text_extracted"
                elif suffix in {".pptx", ".pptm", ".ppt"}:
                    source_format = "pptx_text_extracted"
                elif suffix in {".atom", ".rss", ".xml"}:
                    source_format = "xml_text_extracted"
            else:
                # Content sniffing: handle docs saved without normal suffix.
                try:
                    with real_path.open("rb") as fp:
                        sniff = fp.read(512 * 1024)
                    head = sniff[:8]
                except Exception:
                    sniff = b""
                    head = b""
                if head.startswith(b"%PDF-"):
                    source_format = "pdf_text_extracted"
                    try:
                        full_text = extract_pdf_text_from_path(real_path, max_chars=1_000_000)
                    except Exception as exc:
                        full_text = f"[文档解析失败: {exc}]"
                elif head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
                    from app.attachments import extract_document_text, looks_like_outlook_msg_bytes  # lazy import

                    if looks_like_outlook_msg_bytes(sniff):
                        source_format = "msg_text_extracted"
                        full_text = extract_document_text(str(real_path), max_chars=1_000_000) or ""
                    else:
                        full_text = real_path.read_text(encoding="utf-8", errors="ignore")
                elif head.startswith(b"PK\x03\x04"):
                    from app.attachments import extract_document_text, looks_like_pptx_file, looks_like_xlsx_file  # lazy import

                    if looks_like_xlsx_file(real_path):
                        source_format = "xlsx_text_extracted"
                        full_text = extract_document_text(str(real_path), max_chars=1_000_000) or ""
                    elif looks_like_pptx_file(real_path):
                        source_format = "pptx_text_extracted"
                        full_text = extract_document_text(str(real_path), max_chars=1_000_000) or ""
                    else:
                        full_text = real_path.read_text(encoding="utf-8", errors="ignore")
                else:
                    full_text = real_path.read_text(encoding="utf-8", errors="ignore")

            total_length = len(full_text)
            limit = max(128, min(1_000_000, int(max_chars)))
            start = max(0, int(start_char))
            if start > total_length:
                start = total_length
            end = min(total_length, start + limit)
            text = full_text[start:end]
            truncated = end < total_length
            return {
                "ok": True,
                "path": str(real_path),
                "content": text,
                "length": len(text),
                "start_char": start,
                "end_char": end,
                "total_length": total_length,
                "truncated": truncated,
                "has_more": truncated,
                "source_format": source_format,
            }
        except Exception as exc:
            return {"ok": False, "error": f"read_text_file failed: {exc}"}

    def search_text_in_file(
        self,
        path: str,
        query: str,
        max_matches: int = 8,
        context_chars: int = 280,
    ) -> dict[str, Any]:
        try:
            normalized_query = _normalize_search_query(query)
            if not normalized_query:
                return {"ok": False, "error": "query is empty"}

            variants = _expand_search_variants(normalized_query)
            limit = max(1, min(20, int(max_matches)))
            window = max(40, min(2000, int(context_chars)))
            matches: list[dict[str, Any]] = []

            real_path = _resolve_source_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            if _looks_like_pdf_path(real_path):
                pages = extract_pdf_page_texts_from_path(real_path)
                for variant in variants:
                    pattern = _build_search_pattern(variant)
                    if pattern is None:
                        continue
                    for page_num, body in pages:
                        for found in pattern.finditer(body):
                            span = found.span()
                            start = max(0, span[0] - window)
                            end = min(len(body), span[1] + window)
                            matches.append(
                                {
                                    "query_variant": variant,
                                    "matched_text": found.group(0),
                                    "start_char": span[0],
                                    "end_char": span[1],
                                    "page_hint": page_num,
                                    "context": body[start:end].strip(),
                                    "read_hint": {
                                        "page_hint": page_num,
                                        "start_char": max(0, span[0] - 2000),
                                        "max_chars": 6000,
                                    },
                                }
                            )
                            if len(matches) >= limit:
                                break
                        if len(matches) >= limit:
                            break
                    if len(matches) >= limit:
                        break

                return {
                    "ok": True,
                    "path": str(real_path),
                    "source_format": "pdf_text_extracted",
                    "query": normalized_query,
                    "searched_variants": variants,
                    "match_count": len(matches),
                    "matches": matches,
                    "note": (
                        "Search was run page-by-page over extracted PDF text. "
                        "If match_count=0, only conclude that the current extracted PDF text did not show a hit."
                    ),
                }

            base = self.read_text_file(path=path, start_char=0, max_chars=1_000_000)
            if not bool(base.get("ok")):
                return base

            text = str(base.get("content") or "")
            seen_spans: list[tuple[int, int]] = []
            for variant in variants:
                pattern = _build_search_pattern(variant)
                if pattern is None:
                    continue
                for found in pattern.finditer(text):
                    span = found.span()
                    if any(_spans_overlap(span, prior) for prior in seen_spans):
                        continue
                    seen_spans.append(span)

                    start = max(0, span[0] - window)
                    end = min(len(text), span[1] + window)
                    page_hint = _page_hint_for_offset(text, span[0])
                    matches.append(
                        {
                            "query_variant": variant,
                            "matched_text": found.group(0),
                            "start_char": span[0],
                            "end_char": span[1],
                            "page_hint": page_hint,
                            "context": text[start:end].strip(),
                            "read_hint": {
                                "start_char": max(0, span[0] - 2000),
                                "max_chars": 6000,
                            },
                        }
                    )
                    if len(matches) >= limit:
                        break
                if len(matches) >= limit:
                    break

            return {
                "ok": True,
                "path": str(base.get("path") or real_path),
                "source_format": base.get("source_format") or "text_utf8",
                "query": normalized_query,
                "searched_variants": variants,
                "match_count": len(matches),
                "matches": matches,
                "note": (
                    "Search was run over extracted document text. "
                    "If match_count=0, only conclude that the current extracted text did not show a hit."
                ),
            }
        except Exception as exc:
            return {"ok": False, "error": f"search_text_in_file failed: {exc}"}

    def multi_query_search(
        self,
        path: str,
        queries: list[str],
        per_query_max_matches: int = 3,
        context_chars: int = 280,
    ) -> dict[str, Any]:
        try:
            cleaned_queries = [_normalize_search_query(item) for item in (queries or []) if str(item or "").strip()]
            if not cleaned_queries:
                return {"ok": False, "error": "queries is empty"}

            merged: list[dict[str, Any]] = []
            seen: set[tuple[Any, ...]] = set()
            for query in cleaned_queries[:20]:
                result = self.search_text_in_file(
                    path=path,
                    query=query,
                    max_matches=max(1, min(10, int(per_query_max_matches))),
                    context_chars=context_chars,
                )
                if not bool(result.get("ok")):
                    return result
                for match in result.get("matches") or []:
                    if not isinstance(match, dict):
                        continue
                    key = (
                        match.get("page_hint"),
                        match.get("start_char"),
                        match.get("end_char"),
                        str(match.get("matched_text") or ""),
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    merged.append(match)

            return {
                "ok": True,
                "path": str(_resolve_source_path(self.config, path)),
                "queries": cleaned_queries[:20],
                "match_count": len(merged),
                "matches": merged,
            }
        except Exception as exc:
            return {"ok": False, "error": f"multi_query_search failed: {exc}"}

    def doc_index_build(self, path: str, force_rebuild: bool = False, max_headings: int = 400) -> dict[str, Any]:
        try:
            real_path = _resolve_source_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            if not _looks_like_pdf_path(real_path):
                return {
                    "ok": False,
                    "error": "doc_index_build currently supports PDF files only",
                }

            if force_rebuild:
                clear_pdf_cache_for_path(real_path)
            index = build_pdf_document_index(real_path, force_rebuild=False, max_headings=max_headings)
            headings = index.get("headings") or []
            return {
                "ok": True,
                "path": index.get("path"),
                "cache_path": index.get("cache_path"),
                "cached": bool(index.get("cached")),
                "page_count": int(index.get("page_count") or 0),
                "heading_count": int(index.get("heading_count") or 0),
                "table_page_count": len(index.get("table_pages") or []),
                "headings_preview": headings[:20],
            }
        except Exception as exc:
            return {"ok": False, "error": f"doc_index_build failed: {exc}"}

    def read_section_by_heading(self, path: str, heading: str, max_chars: int = 12000) -> dict[str, Any]:
        try:
            real_path = _resolve_source_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            limit = max(512, min(50000, int(max_chars)))
            if _looks_like_pdf_path(real_path):
                pages = extract_pdf_page_texts_from_path(real_path)
                headings = extract_heading_entries_from_pages(pages, max_headings=1000)
                section = _extract_section_from_pdf_pages(pages, headings, heading, limit)
                if not bool(section.get("ok")):
                    return section
                return {
                    "ok": True,
                    "path": str(real_path),
                    "matched_heading": section.get("matched_heading"),
                    "matched_section": section.get("matched_section"),
                    "page_start": section.get("page_start"),
                    "page_end": section.get("page_end"),
                    "content": section.get("content"),
                }

            base = self.read_text_file(path=path, start_char=0, max_chars=1_000_000)
            if not bool(base.get("ok")):
                return base
            text = str(base.get("content") or "")
            lines = text.splitlines()
            pages = [(1, text)]
            headings = extract_heading_entries_from_pages(pages, max_headings=1000)
            section = _extract_section_from_pdf_pages([(1, text)], headings, heading, limit)
            if bool(section.get("ok")):
                return {
                    "ok": True,
                    "path": str(real_path),
                    "matched_heading": section.get("matched_heading"),
                    "matched_section": section.get("matched_section"),
                    "page_start": 1,
                    "page_end": 1,
                    "content": section.get("content"),
                }
            return {"ok": False, "error": f"Heading not found: {heading}", "path": str(real_path), "line_count": len(lines)}
        except Exception as exc:
            return {"ok": False, "error": f"read_section_by_heading failed: {exc}"}

    def table_extract(
        self,
        path: str,
        query: str = "",
        page_hint: int = 0,
        max_tables: int = 5,
        max_rows: int = 25,
    ) -> dict[str, Any]:
        try:
            real_path = _resolve_source_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            limit_tables = max(1, min(20, int(max_tables)))
            limit_rows = max(1, min(200, int(max_rows)))
            query_norm = _normalize_search_query(query)

            if _looks_like_pdf_path(real_path):
                candidate_pages: list[int] = []
                if page_hint > 0:
                    candidate_pages.append(int(page_hint))
                if query_norm:
                    search = self.search_text_in_file(path=path, query=query_norm, max_matches=8, context_chars=120)
                    if bool(search.get("ok")):
                        candidate_pages.extend(
                            int(item.get("page_hint") or 0)
                            for item in (search.get("matches") or [])
                            if int(item.get("page_hint") or 0) > 0
                        )
                page_numbers = sorted(set(page for page in candidate_pages if page > 0)) or None
                tables = extract_pdf_tables_from_path(
                    real_path,
                    page_numbers=page_numbers,
                    max_tables=limit_tables,
                    max_rows=limit_rows,
                )
                if query_norm:
                    query_tokens = [normalize_lookup_text(query_norm)]
                    filtered: list[dict[str, object]] = []
                    for table in tables:
                        rows = [str(row) for row in table.get("rows") or []]
                        joined = normalize_lookup_text("\n".join(rows))
                        if any(token in joined for token in query_tokens):
                            filtered.append(table)
                    tables = filtered
                return {
                    "ok": True,
                    "path": str(real_path),
                    "table_count": len(tables),
                    "tables": tables[:limit_tables],
                }

            if real_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
                try:
                    from openpyxl import load_workbook  # lazy import
                except Exception as exc:
                    return {"ok": False, "error": f"table_extract requires openpyxl: {exc}"}
                wb = load_workbook(filename=str(real_path), read_only=True, data_only=True)
                try:
                    tables: list[dict[str, Any]] = []
                    for sheet in wb.worksheets:
                        rows: list[str] = []
                        for row in sheet.iter_rows(values_only=True):
                            cells = [_xlsx_cell_to_text(cell) for cell in row]
                            if not any(cells):
                                continue
                            row_line = " | ".join(cells)
                            if query_norm and normalize_lookup_text(query_norm) not in normalize_lookup_text(row_line):
                                continue
                            rows.append(row_line)
                            if len(rows) >= limit_rows:
                                break
                        if rows:
                            tables.append({"sheet": sheet.title or "Sheet", "rows": rows})
                        if len(tables) >= limit_tables:
                            break
                    return {"ok": True, "path": str(real_path), "table_count": len(tables), "tables": tables}
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass

            return {"ok": False, "error": "table_extract currently supports PDF/XLSX files only"}
        except Exception as exc:
            return {"ok": False, "error": f"table_extract failed: {exc}"}

    def fact_check_file(
        self,
        path: str,
        claim: str,
        queries: list[str] | None = None,
        max_evidence: int = 6,
    ) -> dict[str, Any]:
        try:
            cleaned_claim = (claim or "").strip()
            if not cleaned_claim:
                return {"ok": False, "error": "claim is empty"}
            query_list = [_normalize_search_query(item) for item in (queries or []) if str(item or "").strip()]
            if not query_list:
                query_list = _derive_fact_check_queries(cleaned_claim)
            search = self.multi_query_search(
                path=path,
                queries=query_list,
                per_query_max_matches=max(1, min(6, int(max_evidence))),
                context_chars=220,
            )
            if not bool(search.get("ok")):
                return search

            evidence = list(search.get("matches") or [])[: max(1, min(12, int(max_evidence)))]
            verdict = "insufficient_evidence"
            if evidence:
                verdict = "conflicted" if _is_negative_claim(cleaned_claim) else "supported"
            return {
                "ok": True,
                "path": str(_resolve_source_path(self.config, path)),
                "claim": cleaned_claim,
                "queries_used": query_list,
                "verdict": verdict,
                "evidence_count": len(evidence),
                "evidence": evidence,
                "note": (
                    "This tool checks whether the current extracted file text contains evidence related to the claim. "
                    "A 'supported' result still requires agent judgment about relevance and exact wording."
                ),
            }
        except Exception as exc:
            return {"ok": False, "error": f"fact_check_file failed: {exc}"}

    def search_codebase(
        self,
        query: str,
        root: str = ".",
        max_matches: int = 20,
        file_glob: str = "",
        use_regex: bool = False,
        case_sensitive: bool = False,
    ) -> dict[str, Any]:
        try:
            cleaned_query = str(query or "").strip()
            if not cleaned_query:
                return {"ok": False, "error": "query is empty"}
            real_root = _resolve_workspace_path(self.config, root)
            if not real_root.exists():
                return {"ok": False, "error": f"Path not found: {root}"}
            if not real_root.is_dir():
                return {"ok": False, "error": f"Not a directory: {root}"}

            limit = max(1, min(100, int(max_matches)))
            matches: list[dict[str, Any]] = []
            parser_mode = "json"
            if shutil.which("rg"):
                argv_core = ["-n", "--color", "never", "--max-count", str(limit)]
                if not use_regex:
                    argv_core.append("-F")
                if case_sensitive:
                    argv_core.append("-s")
                else:
                    argv_core.append("-i")
                if file_glob.strip():
                    argv_core.extend(["-g", file_glob.strip()])
                argv_tail = [cleaned_query, str(real_root)]

                proc = subprocess.run(["rg", "--json", *argv_core, *argv_tail], capture_output=True, text=True, timeout=20)
                stderr_text = (proc.stderr or "").strip()
                if proc.returncode not in {0, 1} and "--json" in stderr_text.lower():
                    parser_mode = "text_fallback"
                    proc = subprocess.run(
                        ["rg", *argv_core, "--no-heading", *argv_tail],
                        capture_output=True,
                        text=True,
                        timeout=20,
                    )

                if proc.returncode not in {0, 1}:
                    return {"ok": False, "error": (proc.stderr or proc.stdout or "rg failed").strip()}

                if parser_mode == "json":
                    for raw_line in (proc.stdout or "").splitlines():
                        try:
                            event = json.loads(raw_line)
                        except Exception:
                            continue
                        if str(event.get("type") or "") != "match":
                            continue
                        data = event.get("data") if isinstance(event.get("data"), dict) else {}
                        path_block = data.get("path") if isinstance(data.get("path"), dict) else {}
                        lines_block = data.get("lines") if isinstance(data.get("lines"), dict) else {}
                        file_path = str(path_block.get("text") or "").strip()
                        if not file_path:
                            continue
                        try:
                            line_no = int(data.get("line_number") or 0)
                        except Exception:
                            line_no = 0
                        text_line = str(lines_block.get("text") or "").rstrip("\r\n")
                        matches.append(
                            {
                                "path": file_path,
                                "line": line_no,
                                "text": text_line.strip(),
                            }
                        )
                        if len(matches) >= limit:
                            break
                else:
                    for line in (proc.stdout or "").splitlines():
                        parts = line.rsplit(":", 2)
                        if len(parts) != 3:
                            continue
                        file_path, line_no_raw, text_line = parts
                        try:
                            line_no = int(line_no_raw)
                        except Exception:
                            line_no = 0
                        matches.append(
                            {
                                "path": file_path,
                                "line": line_no,
                                "text": text_line.strip(),
                            }
                        )
                        if len(matches) >= limit:
                            break
            else:
                parser_mode = "python_fallback"
                if use_regex:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    pattern = re.compile(cleaned_query, flags)
                else:
                    needle = cleaned_query if case_sensitive else cleaned_query.lower()
                    pattern = None

                for file_path in real_root.rglob("*"):
                    if not file_path.is_file():
                        continue
                    if file_glob.strip():
                        rel = file_path.relative_to(real_root).as_posix()
                        if not fnmatch.fnmatch(rel, file_glob.strip()):
                            continue
                    try:
                        text = file_path.read_text(encoding="utf-8", errors="ignore")
                    except Exception:
                        continue
                    if "\x00" in text:
                        continue
                    for idx, line in enumerate(text.splitlines(), start=1):
                        hay = line if case_sensitive else line.lower()
                        matched = bool(pattern.search(line)) if pattern is not None else needle in hay
                        if not matched:
                            continue
                        matches.append(
                            {
                                "path": str(file_path),
                                "line": idx,
                                "text": line.strip(),
                            }
                        )
                        if len(matches) >= limit:
                            break
                    if len(matches) >= limit:
                        break
            return {
                "ok": True,
                "root": str(real_root),
                "query": cleaned_query,
                "match_count": len(matches),
                "matches": matches,
                "parser_mode": parser_mode,
            }
        except FileNotFoundError:
            return {"ok": False, "error": "rg not found"}
        except Exception as exc:
            return {"ok": False, "error": f"search_codebase failed: {exc}"}

    def copy_file(
        self, src_path: str, dst_path: str, overwrite: bool = True, create_dirs: bool = True
    ) -> dict[str, Any]:
        try:
            src_real = _resolve_source_path(self.config, src_path)
            dst_real = _resolve_workspace_path(self.config, dst_path)

            if not src_real.exists():
                return {"ok": False, "error": f"Source path not found: {src_path}"}
            if not src_real.is_file():
                return {"ok": False, "error": f"Source is not a file: {src_path}"}
            if src_real == dst_real:
                return {"ok": False, "error": "Source and destination are the same file"}

            if dst_real.exists() and dst_real.is_dir():
                return {"ok": False, "error": f"Destination is a directory: {dst_path}"}
            if dst_real.exists() and not overwrite:
                return {"ok": False, "error": f"Destination exists and overwrite=false: {dst_path}"}

            if not dst_real.parent.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Destination parent not found: {dst_real.parent}"}
                dst_real.parent.mkdir(parents=True, exist_ok=True)

            action = "overwrite" if dst_real.exists() else "create"
            shutil.copy2(src_real, dst_real)
            return {
                "ok": True,
                "src_path": str(src_real),
                "dst_path": str(dst_real),
                "action": action,
                "bytes": dst_real.stat().st_size,
            }
        except Exception as exc:
            return {"ok": False, "error": f"copy_file failed: {exc}"}

    def extract_zip(
        self,
        zip_path: str,
        dst_dir: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        max_entries: int = 20000,
        max_total_bytes: int = 524288000,
    ) -> dict[str, Any]:
        try:
            zip_real = _resolve_source_path(self.config, zip_path)
            if not zip_real.exists():
                return {"ok": False, "error": f"Zip path not found: {zip_path}"}
            if not zip_real.is_file():
                return {"ok": False, "error": f"Zip path is not a file: {zip_path}"}

            dst_raw = (dst_dir or "").strip()
            if not dst_raw:
                dst_raw = str(zip_real.with_suffix(""))

            dst_real = _resolve_workspace_path(self.config, dst_raw)
            if dst_real.exists() and dst_real.is_file():
                return {"ok": False, "error": f"Destination is a file, not directory: {dst_raw}"}
            if not dst_real.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Destination directory not found: {dst_real}"}
                dst_real.mkdir(parents=True, exist_ok=True)

            entry_limit = max(1, min(100000, int(max_entries)))
            total_limit = max(1024, min(2147483648, int(max_total_bytes)))

            extracted_files = 0
            skipped_files = 0
            extracted_bytes = 0

            with zipfile.ZipFile(zip_real, "r") as zf:
                infos = zf.infolist()
                if len(infos) > entry_limit:
                    return {
                        "ok": False,
                        "error": f"Zip entries exceed max_entries limit ({len(infos)} > {entry_limit}).",
                    }

                total_uncompressed = sum(int(getattr(i, "file_size", 0) or 0) for i in infos)
                if total_uncompressed > total_limit:
                    return {
                        "ok": False,
                        "error": (
                            f"Zip uncompressed size exceeds max_total_bytes "
                            f"({total_uncompressed} > {total_limit})."
                        ),
                    }

                for info in infos:
                    name = (info.filename or "").replace("\\", "/")
                    if not name:
                        continue

                    rel = Path(name)
                    if rel.is_absolute() or ".." in rel.parts:
                        return {"ok": False, "error": f"Unsafe zip entry path detected: {name}"}

                    target = (dst_real / rel).resolve()
                    if not _is_within(target, dst_real):
                        return {"ok": False, "error": f"Unsafe zip entry path detected: {name}"}

                    if info.is_dir():
                        target.mkdir(parents=True, exist_ok=True)
                        continue

                    if target.exists() and not overwrite:
                        skipped_files += 1
                        continue

                    target.parent.mkdir(parents=True, exist_ok=True)
                    with zf.open(info, "r") as src, open(target, "wb") as out:
                        shutil.copyfileobj(src, out)
                    extracted_files += 1
                    extracted_bytes += int(target.stat().st_size)

            return {
                "ok": True,
                "zip_path": str(zip_real),
                "dst_dir": str(dst_real),
                "files_extracted": extracted_files,
                "files_skipped": skipped_files,
                "bytes_extracted": extracted_bytes,
                "overwrite": bool(overwrite),
            }
        except zipfile.BadZipFile:
            return {"ok": False, "error": f"Invalid zip archive: {zip_path}"}
        except Exception as exc:
            return {"ok": False, "error": f"extract_zip failed: {exc}"}

    def extract_msg_attachments(
        self,
        msg_path: str,
        dst_dir: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        max_attachments: int = 500,
        max_total_bytes: int = 524288000,
    ) -> dict[str, Any]:
        try:
            msg_real = _resolve_source_path(self.config, msg_path)
            if not msg_real.exists():
                return {"ok": False, "error": f"MSG path not found: {msg_path}"}
            if not msg_real.is_file():
                return {"ok": False, "error": f"MSG path is not a file: {msg_path}"}

            from app.attachments import looks_like_outlook_msg_file  # lazy import

            suffix = msg_real.suffix.lower()
            if suffix != ".msg" and not looks_like_outlook_msg_file(msg_real):
                return {"ok": False, "error": f"Not an Outlook .msg file: {msg_path}"}

            dst_raw = (dst_dir or "").strip()
            if not dst_raw:
                dst_raw = str(msg_real.parent / f"{msg_real.stem}_attachments")

            dst_real = _resolve_workspace_path(self.config, dst_raw)
            if dst_real.exists() and dst_real.is_file():
                return {"ok": False, "error": f"Destination is a file, not directory: {dst_raw}"}
            if not dst_real.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Destination directory not found: {dst_real}"}
                dst_real.mkdir(parents=True, exist_ok=True)

            attachment_limit = max(1, min(5000, int(max_attachments)))
            total_limit = max(1024, min(2147483648, int(max_total_bytes)))

            try:
                import extract_msg  # lazy import
            except Exception as exc:
                return {
                    "ok": False,
                    "error": (
                        "解析 .msg 附件需要依赖 extract-msg。请执行 "
                        "`pip install -r requirements.txt` 后重试。"
                    ),
                    "detail": str(exc),
                }

            msg = extract_msg.openMsg(str(msg_real), strict=False, delayAttachments=False)
            try:
                attachments = list(getattr(msg, "attachments", []) or [])
                if len(attachments) > attachment_limit:
                    return {
                        "ok": False,
                        "error": (
                            f"MSG attachments exceed max_attachments limit "
                            f"({len(attachments)} > {attachment_limit})."
                        ),
                        "msg_path": str(msg_real),
                        "dst_dir": str(dst_real),
                    }

                entries: list[dict[str, Any]] = []
                files_saved = 0
                files_skipped = 0
                bytes_extracted = 0

                for idx, att in enumerate(attachments, start=1):
                    raw_name = (
                        (getattr(att, "longFilename", None) or "")
                        or (getattr(att, "filename", None) or "")
                        or (getattr(att, "name", None) or "")
                        or f"attachment_{idx}"
                    )
                    safe_name = _safe_filename(str(raw_name or ""))
                    if safe_name == "download.bin":
                        safe_name = f"attachment_{idx}.bin"

                    att_type = str(getattr(att, "type", "") or "").upper()
                    if "MSG" in att_type and not safe_name.lower().endswith(".msg"):
                        safe_name = f"{safe_name}.msg"

                    target = (dst_real / safe_name).resolve()
                    if not _is_within(target, dst_real):
                        return {
                            "ok": False,
                            "error": f"Unsafe attachment path detected: {safe_name}",
                            "msg_path": str(msg_real),
                            "dst_dir": str(dst_real),
                        }

                    if target.exists() and not overwrite:
                        files_skipped += 1
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "skipped_exists",
                                "path": str(target),
                            }
                        )
                        continue

                    try:
                        save_result = att.save(
                            customPath=str(dst_real),
                            customFilename=safe_name,
                            overwriteExisting=bool(overwrite),
                            extractEmbedded=True,
                            skipEmbedded=False,
                        )
                    except Exception as exc:
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "error",
                                "error": str(exc),
                            }
                        )
                        continue

                    saved_paths: list[Path] = []
                    if (
                        isinstance(save_result, tuple)
                        and len(save_result) >= 2
                        and save_result[1] is not None
                    ):
                        payload = save_result[1]
                        if isinstance(payload, str):
                            saved_paths.append(Path(payload).resolve())
                        elif isinstance(payload, list):
                            for item in payload:
                                if isinstance(item, str):
                                    saved_paths.append(Path(item).resolve())

                    if not saved_paths and target.exists():
                        saved_paths.append(target)

                    saved_payload: list[dict[str, Any]] = []
                    for path_obj in saved_paths:
                        if not path_obj.exists():
                            continue
                        if not _is_within(path_obj, dst_real):
                            continue
                        size = path_obj.stat().st_size if path_obj.is_file() else None
                        if isinstance(size, int):
                            bytes_extracted += size
                        saved_payload.append(
                            {
                                "path": str(path_obj),
                                "is_dir": path_obj.is_dir(),
                                "bytes": size,
                            }
                        )

                    if bytes_extracted > total_limit:
                        return {
                            "ok": False,
                            "error": (
                                f"Extracted bytes exceed max_total_bytes limit "
                                f"({bytes_extracted} > {total_limit})."
                            ),
                            "msg_path": str(msg_real),
                            "dst_dir": str(dst_real),
                            "attachments_total": len(attachments),
                            "files_saved": files_saved,
                            "files_skipped": files_skipped,
                            "bytes_extracted": bytes_extracted,
                            "entries": entries,
                        }

                    if saved_payload:
                        files_saved += 1
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "saved",
                                "saved": saved_payload,
                            }
                        )
                    else:
                        entries.append(
                            {
                                "index": idx,
                                "name": safe_name,
                                "status": "no_output",
                            }
                        )

                return {
                    "ok": True,
                    "msg_path": str(msg_real),
                    "dst_dir": str(dst_real),
                    "attachments_total": len(attachments),
                    "files_saved": files_saved,
                    "files_skipped": files_skipped,
                    "bytes_extracted": bytes_extracted,
                    "entries": entries,
                    "overwrite": bool(overwrite),
                }
            finally:
                close = getattr(msg, "close", None)
                if callable(close):
                    try:
                        close()
                    except Exception:
                        pass
        except Exception as exc:
            return {"ok": False, "error": f"extract_msg_attachments failed: {exc}"}

    def write_text_file(
        self, path: str, content: str, overwrite: bool = True, create_dirs: bool = True
    ) -> dict[str, Any]:
        try:
            real_path = _resolve_workspace_path(self.config, path)
            if real_path.exists() and real_path.is_dir():
                return {"ok": False, "error": f"Path is a directory, not a file: {path}"}

            if real_path.exists() and not overwrite:
                return {"ok": False, "error": f"File already exists and overwrite=false: {path}"}

            if not real_path.parent.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Parent directory not found: {real_path.parent}"}
                real_path.parent.mkdir(parents=True, exist_ok=True)

            action = "overwrite" if real_path.exists() else "create"
            real_path.write_text(content, encoding="utf-8")
            return {
                "ok": True,
                "path": str(real_path),
                "action": action,
                "chars": len(content),
                "bytes_utf8": len(content.encode("utf-8")),
            }
        except Exception as exc:
            return {"ok": False, "error": f"write_text_file failed: {exc}"}

    def append_text_file(
        self,
        path: str,
        content: str,
        create_if_missing: bool = True,
        create_dirs: bool = True,
    ) -> dict[str, Any]:
        try:
            real_path = _resolve_workspace_path(self.config, path)
            if real_path.exists() and real_path.is_dir():
                return {"ok": False, "error": f"Path is a directory, not a file: {path}"}
            if not real_path.exists() and not create_if_missing:
                return {"ok": False, "error": f"File not found and create_if_missing=false: {path}"}

            if not real_path.parent.exists():
                if not create_dirs:
                    return {"ok": False, "error": f"Parent directory not found: {real_path.parent}"}
                real_path.parent.mkdir(parents=True, exist_ok=True)

            created = not real_path.exists()
            with real_path.open("a", encoding="utf-8") as fp:
                fp.write(content)
            return {
                "ok": True,
                "path": str(real_path),
                "action": "create" if created else "append",
                "chars_appended": len(content),
                "bytes_appended_utf8": len(content.encode("utf-8")),
                "bytes_total_utf8": real_path.stat().st_size,
            }
        except Exception as exc:
            return {"ok": False, "error": f"append_text_file failed: {exc}"}

    def replace_in_file(
        self,
        path: str,
        old_text: str,
        new_text: str,
        replace_all: bool = False,
        max_replacements: int = 1,
    ) -> dict[str, Any]:
        if not old_text:
            return {"ok": False, "error": "old_text cannot be empty"}
        if max_replacements < 1:
            return {"ok": False, "error": "max_replacements must be >= 1"}

        try:
            real_path = _resolve_workspace_path(self.config, path)
            if not real_path.exists():
                return {"ok": False, "error": f"Path not found: {path}"}
            if not real_path.is_file():
                return {"ok": False, "error": f"Not a file: {path}"}

            source = real_path.read_text(encoding="utf-8", errors="ignore")
            found = source.count(old_text)
            if found <= 0:
                return {"ok": False, "error": "Target text not found", "path": str(real_path)}

            limit = found if replace_all else min(found, max(1, min(200, max_replacements)))
            updated = source.replace(old_text, new_text, limit)
            real_path.write_text(updated, encoding="utf-8")
            return {
                "ok": True,
                "path": str(real_path),
                "replacements": limit,
                "remaining_matches": max(0, found - limit),
                "chars": len(updated),
                "bytes_utf8": len(updated.encode("utf-8")),
            }
        except Exception as exc:
            return {"ok": False, "error": f"replace_in_file failed: {exc}"}

    def _domain_allowed(self, host: str) -> bool:
        if self.config.web_allow_all_domains:
            return True

        host = host.lower().strip(".")
        for allowed in self.config.web_allowed_domains:
            d = allowed.lower().strip(".")
            if host == d or host.endswith("." + d):
                return True
        return False

    def search_web(self, query: str, max_results: int = 5, timeout_sec: int = 12) -> dict[str, Any]:
        q = (query or "").strip()
        if not q:
            return {"ok": False, "error": "query cannot be empty"}

        timeout_val = max(3, min(30, timeout_sec))
        limit = max(1, min(20, int(max_results)))
        cache_key = {"query": q, "max_results": limit, "algo_version": 3}
        cached = self._load_web_cache("search_web", cache_key, max_age_sec=900)
        if cached:
            return {**cached, "cached": True}
        read_limit = min(500000, max(20000, self.config.web_fetch_max_chars))
        ddg_allowed = self._domain_allowed("duckduckgo.com")
        prefer_news = _looks_news_like_query(q)
        prefer_baseball = _looks_baseball_query(q)
        query_is_specific = _query_looks_specific(q)
        rss_candidates = _build_rss_candidates(q)
        rss_allowed_candidates: list[tuple[str, str]] = []
        for name, url in rss_candidates:
            host = (urllib.parse.urlsplit(url).hostname or "").strip().lower()
            if host and self._domain_allowed(host):
                rss_allowed_candidates.append((name, url))

        if not ddg_allowed and not rss_allowed_candidates:
            return {
                "ok": False,
                "error": (
                    "Domain not allowed for search engines and RSS sources. "
                    f"Allowed: {', '.join(self.config.web_allowed_domains)}"
                ),
            }

        search_url = "https://duckduckgo.com/html/?q=" + urllib.parse.quote_plus(q)
        lite_url = "https://lite.duckduckgo.com/lite/?q=" + urllib.parse.quote_plus(q)

        if self.config.web_skip_tls_verify:
            ssl_context = ssl._create_unverified_context()
        elif self.config.web_ca_cert_path:
            try:
                ssl_context = ssl.create_default_context(cafile=self.config.web_ca_cert_path)
            except Exception as exc:
                return {
                    "ok": False,
                    "error": f"Invalid web CA cert path: {self.config.web_ca_cert_path} ({exc})",
                }
        else:
            ssl_context = ssl.create_default_context()

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.5",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        tls_warning: str | None = None
        active_context = ssl_context

        def _open(current_context: ssl.SSLContext | None, target_url: str):
            req = urllib.request.Request(
                url=target_url,
                headers=headers,
                method="GET",
            )
            opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=current_context))
            return opener.open(req, timeout=timeout_val)

        def _fetch_page(target_url: str, current_context: ssl.SSLContext | None) -> tuple[int, str, str, bool]:
            with _open(current_context, target_url) as resp:
                status = getattr(resp, "status", None) or 200
                content_type = (resp.headers.get("Content-Type") or "").lower()
                raw = resp.read(read_limit + 1)
                truncated = len(raw) > read_limit
                raw = raw[:read_limit]
                text = raw.decode("utf-8", errors="ignore")
                return status, content_type, text, truncated

        def _fetch_page_with_retry(target_url: str) -> tuple[int, str, str, bool]:
            nonlocal active_context, tls_warning
            try:
                return _fetch_page(target_url, active_context)
            except Exception as first_exc:
                if not self.config.web_skip_tls_verify and _is_cert_verify_error(first_exc):
                    tls_warning = "TLS verify failed; search_web auto-retried with verify disabled."
                    active_context = ssl._create_unverified_context()
                    return _fetch_page(target_url, active_context)
                raise

        try:
            results: list[dict[str, str]] = []
            source = "unknown"
            status = 200
            content_type = "text/html"
            truncated = False
            warning_parts: list[str] = []
            seen_result_keys: set[str] = set()

            def _append_results(items: list[dict[str, str]], source_name: str) -> int:
                added = 0
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    title = str(item.get("title", "")).strip()
                    url = str(item.get("url", "")).strip()
                    if not title and not url:
                        continue
                    key = f"{title}|{url}".lower()
                    if key in seen_result_keys:
                        continue
                    seen_result_keys.add(key)

                    row = dict(item)
                    row.setdefault("source", source_name)
                    results.append(row)
                    added += 1
                    if len(results) >= limit:
                        break
                return added

            def _rss_source_requires_query_match(source_name: str) -> bool:
                return query_is_specific and source_name in {
                    "mlb_official_rss",
                    "espn_mlb_rss",
                    "yahoo_mlb_rss",
                    "nhk_sports_rss",
                }

            if prefer_news and rss_allowed_candidates:
                for rss_name, rss_url in rss_allowed_candidates:
                    if len(results) >= limit:
                        break
                    try:
                        status, content_type, text, truncated = _fetch_page_with_retry(rss_url)
                        rss_results = _extract_google_news_rss_results(text, max_results=limit)
                        if _rss_source_requires_query_match(rss_name):
                            rss_results = [
                                row for row in rss_results if _query_relevance_score(q, row) >= 12.0
                            ]
                        if _append_results(rss_results, rss_name) > 0 and source == "unknown":
                            source = f"rss:{rss_name}"
                    except Exception as exc:
                        warning_parts.append(f"{rss_name} 获取失败: {exc}")

            ddg_error: str | None = None
            if ddg_allowed and not results:
                try:
                    status, content_type, text, truncated = _fetch_page_with_retry(search_url)
                    ddg_results = _extract_ddg_results(text, max_results=limit)
                    if _append_results(ddg_results, "duckduckgo_html") > 0:
                        source = "duckduckgo_html"
                    if not results:
                        status, content_type, text, truncated = _fetch_page_with_retry(lite_url)
                        ddg_results = _extract_ddg_results(text, max_results=limit)
                        if _append_results(ddg_results, "duckduckgo_lite") > 0:
                            source = "duckduckgo_lite"
                except Exception as exc:
                    ddg_error = str(exc)

            if ddg_error:
                warning_parts.append(f"DuckDuckGo 搜索失败: {ddg_error}")

            if not results and rss_allowed_candidates and not prefer_news:
                for rss_name, rss_url in rss_allowed_candidates:
                    if len(results) >= limit:
                        break
                    try:
                        status, content_type, text, truncated = _fetch_page_with_retry(rss_url)
                        rss_results = _extract_google_news_rss_results(text, max_results=limit)
                        if _rss_source_requires_query_match(rss_name):
                            rss_results = [
                                row for row in rss_results if _query_relevance_score(q, row) >= 12.0
                            ]
                        if _append_results(rss_results, rss_name) > 0 and source == "unknown":
                            source = f"rss:{rss_name}"
                    except Exception as exc:
                        warning_parts.append(f"{rss_name} 回退失败: {exc}")

            if not results and prefer_baseball:
                curated = [
                    {
                        "title": "MLB News (Official)",
                        "url": "https://www.mlb.com/news",
                        "snippet": "Fallback source when search engines are blocked.",
                        "source": "fallback_static",
                    },
                    {
                        "title": "ESPN MLB",
                        "url": "https://www.espn.com/mlb/",
                        "snippet": "Fallback source when search engines are blocked.",
                        "source": "fallback_static",
                    },
                    {
                        "title": "Yahoo Sports MLB",
                        "url": "https://sports.yahoo.com/mlb/",
                        "snippet": "Fallback source when search engines are blocked.",
                        "source": "fallback_static",
                    },
                    {
                        "title": "NPB Official",
                        "url": "https://npb.jp/",
                        "snippet": "Fallback source when search engines are blocked.",
                        "source": "fallback_static",
                    },
                    {
                        "title": "Yahoo Japan NPB",
                        "url": "https://baseball.yahoo.co.jp/npb/",
                        "snippet": "Fallback source when search engines are blocked.",
                        "source": "fallback_static",
                    },
                ]
                for item in curated:
                    host = (urllib.parse.urlsplit(item["url"]).hostname or "").strip().lower()
                    if host and self._domain_allowed(host):
                        results.append(item)
                    if len(results) >= limit:
                        break
                if results:
                    source = "fallback:baseball_static_links"
                    warning_parts.append("实时新闻抓取受限，已回退到可访问的棒球新闻入口链接。")

            if not results:
                warning_parts.append("搜索结果页解析为空，可能被网关改写或反爬。")

            if tls_warning:
                warning_parts.insert(0, tls_warning)

            warning = " ".join(part.strip() for part in warning_parts if part and part.strip()) or None
            if source == "unknown":
                source = "none"

            normalized_results: list[dict[str, Any]] = []
            for item in results:
                if not isinstance(item, dict):
                    continue
                row = dict(item)
                url = str(row.get("url") or "").strip()
                try:
                    row["domain"] = (urllib.parse.urlsplit(url).hostname or "").strip().lower()
                except Exception:
                    row["domain"] = ""
                row["score"] = round(_score_web_result(q, row), 3)
                normalized_results.append(row)
            normalized_results.sort(
                key=lambda item: (
                    float(item.get("score") or 0.0),
                    bool(item.get("published_at")),
                    len(str(item.get("title") or "")),
                ),
                reverse=True,
            )
            if query_is_specific and normalized_results:
                best_score = float(normalized_results[0].get("score") or 0.0)
                if best_score >= 12.0:
                    min_score = max(6.0, best_score * 0.45)
                    normalized_results = [
                        item for item in normalized_results if float(item.get("score") or 0.0) >= min_score
                    ]

            payload = {
                "ok": True,
                "query": q,
                "engine": source,
                "status": status,
                "content_type": content_type,
                "count": len(normalized_results),
                "results": normalized_results,
                "truncated": truncated,
                "warning": warning,
                "cached": False,
            }
            self._save_web_cache("search_web", cache_key, payload)
            return payload
        except urllib.error.HTTPError as exc:
            body = exc.read(4000).decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
            return {"ok": False, "error": f"HTTP {exc.code}: {exc.reason}", "body_preview": body}
        except Exception as exc:
            return {"ok": False, "error": f"search_web failed: {exc}"}

    def download_web_file(
        self,
        url: str,
        dst_path: str = "",
        overwrite: bool = True,
        create_dirs: bool = True,
        timeout_sec: int = 20,
        max_bytes: int = 52_428_800,
    ) -> dict[str, Any]:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            return {"ok": False, "error": "Only http/https URLs are supported"}
        if not parsed.netloc:
            return {"ok": False, "error": "Invalid URL"}

        try:
            request_url = _normalize_url_for_request(url)
        except Exception as exc:
            return {"ok": False, "error": f"Invalid URL: {exc}"}

        host = parsed.hostname or ""
        if not self._domain_allowed(host):
            return {
                "ok": False,
                "error": f"Domain not allowed: {host}. Allowed: {', '.join(self.config.web_allowed_domains)}",
            }

        timeout_val = max(3, min(120, int(timeout_sec)))
        byte_limit = max(1024, min(209_715_200, int(max_bytes)))

        ssl_context: ssl.SSLContext | None = None
        if parsed.scheme == "https":
            if self.config.web_skip_tls_verify:
                ssl_context = ssl._create_unverified_context()
            elif self.config.web_ca_cert_path:
                try:
                    ssl_context = ssl.create_default_context(cafile=self.config.web_ca_cert_path)
                except Exception as exc:
                    return {
                        "ok": False,
                        "error": f"Invalid web CA cert path: {self.config.web_ca_cert_path} ({exc})",
                    }
            else:
                ssl_context = ssl.create_default_context()

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "*/*",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        tls_warning: str | None = None

        def _open(current_context: ssl.SSLContext | None):
            req = urllib.request.Request(
                url=request_url,
                headers=headers,
                method="GET",
            )
            opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=current_context))
            return opener.open(req, timeout=timeout_val)

        try:
            try:
                resp_cm = _open(ssl_context)
            except Exception as first_exc:
                if not self.config.web_skip_tls_verify and _is_cert_verify_error(first_exc):
                    tls_warning = "TLS verify failed; download_web_file auto-retried with verify disabled."
                    resp_cm = _open(ssl._create_unverified_context())
                else:
                    raise

            with resp_cm as resp:
                status = getattr(resp, "status", None) or 200
                content_type = (resp.headers.get("Content-Type") or "").lower()
                content_disposition = resp.headers.get("Content-Disposition") or ""
                filename = _guess_filename_from_response(url=url, content_type=content_type, content_disposition=content_disposition)

                raw = resp.read(byte_limit + 1)
                truncated = len(raw) > byte_limit
                if truncated:
                    return {
                        "ok": False,
                        "error": f"Remote file exceeds max_bytes limit ({byte_limit} bytes).",
                        "status": status,
                        "url": url,
                        "content_type": content_type,
                        "filename": filename,
                    }

                target_raw = (dst_path or "").strip()
                if not target_raw:
                    target_raw = str(Path("downloads") / filename)

                target_path = _resolve_workspace_path(self.config, target_raw)
                if target_path.exists() and target_path.is_dir():
                    return {"ok": False, "error": f"Destination is a directory: {target_raw}"}
                if target_path.exists() and not overwrite:
                    return {"ok": False, "error": f"Destination exists and overwrite=false: {target_raw}"}
                if not target_path.parent.exists():
                    if not create_dirs:
                        return {"ok": False, "error": f"Destination parent not found: {target_path.parent}"}
                    target_path.parent.mkdir(parents=True, exist_ok=True)

                action = "overwrite" if target_path.exists() else "create"
                target_path.write_bytes(raw)
                return {
                    "ok": True,
                    "url": url,
                    "status": status,
                    "content_type": content_type,
                    "path": str(target_path),
                    "bytes": len(raw),
                    "filename": filename,
                    "action": action,
                    "warning": tls_warning,
                }
        except urllib.error.HTTPError as exc:
            body = exc.read(4000).decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
            return {"ok": False, "error": f"HTTP {exc.code}: {exc.reason}", "body_preview": body}
        except Exception as exc:
            return {"ok": False, "error": f"download_web_file failed: {exc}"}

    def fetch_web(self, url: str, max_chars: int = 120000, timeout_sec: int = 12) -> dict[str, Any]:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            return {"ok": False, "error": "Only http/https URLs are supported"}
        if not parsed.netloc:
            return {"ok": False, "error": "Invalid URL"}

        try:
            request_url = _normalize_url_for_request(url)
        except Exception as exc:
            return {"ok": False, "error": f"Invalid URL: {exc}"}

        host = parsed.hostname or ""
        if not self._domain_allowed(host):
            return {
                "ok": False,
                "error": f"Domain not allowed: {host}. Allowed: {', '.join(self.config.web_allowed_domains)}",
            }

        timeout_val = max(3, min(30, timeout_sec))
        limit = max(512, min(500000, max_chars, self.config.web_fetch_max_chars))
        cache_key = {"url": request_url, "max_chars": limit}
        cached = self._load_web_cache("fetch_web", cache_key, max_age_sec=900)
        if cached:
            return {**cached, "cached": True}
        ssl_context: ssl.SSLContext | None = None
        if parsed.scheme == "https":
            if self.config.web_skip_tls_verify:
                ssl_context = ssl._create_unverified_context()
            elif self.config.web_ca_cert_path:
                try:
                    ssl_context = ssl.create_default_context(cafile=self.config.web_ca_cert_path)
                except Exception as exc:
                    return {
                        "ok": False,
                        "error": f"Invalid web CA cert path: {self.config.web_ca_cert_path} ({exc})",
                    }
            else:
                ssl_context = ssl.create_default_context()

        default_headers = {
            # Use a browser-like UA to reduce bot-block false positives.
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/json,text/plain,application/xml;q=0.9,*/*;q=0.5",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        tls_warning: str | None = None

        def _open(current_context: ssl.SSLContext | None, target_url: str):
            req = urllib.request.Request(
                url=target_url,
                headers=default_headers,
                method="GET",
            )
            opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=current_context))
            return opener.open(req, timeout=timeout_val)

        try:
            try:
                resp_cm = _open(ssl_context, request_url)
            except Exception as first_exc:
                # Pragmatic fallback for enterprise TLS chains:
                # if verification fails and user did not explicitly disable it,
                # retry once with verification off for fetch_web only.
                if not self.config.web_skip_tls_verify and _is_cert_verify_error(first_exc):
                    tls_warning = "TLS verify failed; fetch_web auto-retried with verify disabled."
                    resp_cm = _open(ssl._create_unverified_context(), request_url)
                else:
                    raise

            with resp_cm as resp:
                status = getattr(resp, "status", None) or 200
                content_type = (resp.headers.get("Content-Type") or "").lower()
                content_disposition = (resp.headers.get("Content-Disposition") or "").lower()
                pdf_like = (
                    "application/pdf" in content_type
                    or parsed.path.lower().endswith(".pdf")
                    or ".pdf" in content_disposition
                )
                pdf_byte_limit = min(20_000_000, max(1_000_000, self.config.web_fetch_max_chars * 40))
                raw_limit = pdf_byte_limit if pdf_like else limit

                raw = resp.read(raw_limit + 1)
                truncated = len(raw) > raw_limit
                raw = raw[:raw_limit]

                if pdf_like:
                    try:
                        pdf_text = _extract_pdf_text_from_bytes(raw, max_chars=limit)
                        warning = tls_warning
                        if truncated:
                            warning = (
                                f"{warning} PDF 文件较大，已按 {raw_limit} bytes 截断读取。"
                                if warning
                                else f"PDF 文件较大，已按 {raw_limit} bytes 截断读取。"
                            )
                        if not pdf_text.strip():
                            warning = (
                                f"{warning} PDF 可读文本为空（可能是扫描件图片）。"
                                if warning
                                else "PDF 可读文本为空（可能是扫描件图片）。"
                            )
                        payload = {
                            "ok": True,
                            "url": url,
                            "status": status,
                            "content_type": content_type,
                            "domain": host,
                            "binary": False,
                            "truncated": truncated,
                            "content": pdf_text,
                            "length": len(pdf_text),
                            "source_format": "pdf_text_extracted",
                            "warning": warning,
                            "cached": False,
                        }
                        self._save_web_cache("fetch_web", cache_key, payload)
                        return payload
                    except Exception as pdf_exc:
                        warning = (
                            f"{tls_warning} PDF 文本提取失败: {pdf_exc}"
                            if tls_warning
                            else f"PDF 文本提取失败: {pdf_exc}"
                        )
                        return {
                            "ok": True,
                            "url": url,
                            "status": status,
                            "content_type": content_type,
                            "binary": True,
                            "size_preview_bytes": len(raw),
                            "truncated": truncated,
                            "warning": warning,
                        }

                if any(x in content_type for x in ["application/octet-stream", "image/", "audio/", "video/"]):
                    return {
                        "ok": True,
                        "url": url,
                        "status": status,
                        "content_type": content_type,
                        "binary": True,
                        "size_preview_bytes": len(raw),
                        "truncated": truncated,
                        "warning": tls_warning,
                    }

                text = raw.decode("utf-8", errors="ignore")
                if _looks_like_html(content_type, text):
                    metadata = _extract_html_metadata(text, base_url=url)
                    extracted = _extract_html_text(text, max_chars=limit)
                    warning = None
                    if len(extracted.strip()) < 80:
                        warning = (
                            "页面正文较少，可能是 JS 动态渲染或反爬页面。"
                            "建议改用该站点公开 API，或换一个可直读正文的页面。"
                        )
                    if _looks_like_script_payload(extracted):
                        script_warning = (
                            "抓取内容疑似脚本/反爬响应，而非正文页面。"
                            "请不要据此下结论，建议改用官方 API 或可直读页面。"
                        )
                        warning = f"{script_warning} {warning}" if warning else script_warning

                        # Search-engine anti-bot fallback: try a text-friendly results page.
                        search_query = _extract_search_query(url)
                        if search_query and self._domain_allowed("duckduckgo.com"):
                            fallback_url = (
                                "https://duckduckgo.com/html/?q="
                                + urllib.parse.quote_plus(search_query)
                            )
                            try:
                                with _open(ssl_context, fallback_url) as fb_resp:
                                    fb_status = getattr(fb_resp, "status", None) or 200
                                    fb_ct = (fb_resp.headers.get("Content-Type") or "").lower()
                                    fb_raw = fb_resp.read(limit + 1)
                                    fb_truncated = len(fb_raw) > limit
                                    fb_raw = fb_raw[:limit]
                                    fb_text = fb_raw.decode("utf-8", errors="ignore")
                                    fb_extracted = _extract_html_text(fb_text, max_chars=limit)

                                if fb_extracted.strip() and not _looks_like_script_payload(fb_extracted):
                                    if tls_warning:
                                        warning = f"{tls_warning} {warning}" if warning else tls_warning
                                    fallback_warning = (
                                        f"{warning} 已自动回退到 DuckDuckGo HTML 结果页（query={search_query}）。"
                                        if warning
                                        else f"已自动回退到 DuckDuckGo HTML 结果页（query={search_query}）。"
                                    )
                                    fallback_payload = {
                                        "ok": True,
                                        "url": url,
                                        "status": fb_status,
                                        "content_type": fb_ct,
                                        "domain": host,
                                        "binary": False,
                                        "truncated": fb_truncated,
                                        "content": fb_extracted,
                                        "length": len(fb_extracted),
                                        "source_format": "search_fallback_duckduckgo_html",
                                        "warning": fallback_warning,
                                        "title": metadata.get("title") or "",
                                        "published_at": metadata.get("published_at") or "",
                                        "canonical_url": metadata.get("canonical_url") or "",
                                        "cached": False,
                                    }
                                    self._save_web_cache("fetch_web", cache_key, fallback_payload)
                                    return fallback_payload
                            except Exception as fb_exc:
                                warning = (
                                    f"{warning} DuckDuckGo 回退失败: {fb_exc}"
                                    if warning
                                    else f"DuckDuckGo 回退失败: {fb_exc}"
                                )

                        # Avoid passing noisy script payload to the model.
                        extracted = (
                            "[抓取到脚本/反爬页面，正文不可用。"
                            "请改用目标站点公开 API、可直读正文 URL，或非搜索结果页链接。]"
                        )
                    if tls_warning:
                        warning = f"{tls_warning} {warning}" if warning else tls_warning
                    payload = {
                        "ok": True,
                        "url": url,
                        "status": status,
                        "content_type": content_type,
                        "domain": host,
                        "binary": False,
                        "truncated": truncated,
                        "content": extracted,
                        "length": len(extracted),
                        "source_format": "html_text_extracted",
                        "warning": warning,
                        "title": metadata.get("title") or "",
                        "published_at": metadata.get("published_at") or "",
                        "canonical_url": metadata.get("canonical_url") or "",
                        "cached": False,
                    }
                    self._save_web_cache("fetch_web", cache_key, payload)
                    return payload

                payload = {
                    "ok": True,
                    "url": url,
                    "status": status,
                    "content_type": content_type,
                    "domain": host,
                    "binary": False,
                    "truncated": truncated,
                    "content": text,
                    "length": len(text),
                    "warning": tls_warning,
                    "cached": False,
                }
                self._save_web_cache("fetch_web", cache_key, payload)
                return payload
        except urllib.error.HTTPError as exc:
            body = exc.read(4000).decode("utf-8", errors="ignore") if hasattr(exc, "read") else ""
            return {"ok": False, "error": f"HTTP {exc.code}: {exc.reason}", "body_preview": body}
        except Exception as exc:
            return {"ok": False, "error": f"fetch_web failed: {exc}"}


def parse_json_arguments(raw: str | dict[str, Any] | None) -> dict[str, Any]:
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    if not raw.strip():
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}
