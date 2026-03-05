from __future__ import annotations

import base64
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from html import unescape
from pathlib import Path

from app.document_text import extract_pdf_text_from_path, truncate_text

_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_MSG_MARKERS_ASCII = (b"__substg1.0_", b"IPM.")
_MSG_MARKERS_UTF16 = tuple(marker.decode("ascii").encode("utf-16-le") for marker in _MSG_MARKERS_ASCII)
_XLSX_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_PPTX_SUFFIXES = {".pptx", ".pptm"}


def _read_plain_text(path: Path, max_chars: int) -> str:
    text = path.read_text(encoding="utf-8", errors="ignore")
    return truncate_text(text, max_chars)


def _extract_pdf(path: Path, max_chars: int) -> str:
    return extract_pdf_text_from_path(path, max_chars=max_chars)


def _extract_docx(path: Path, max_chars: int) -> str:
    from docx import Document  # lazy import

    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return truncate_text(text, max_chars)


def _xlsx_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        try:
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
        except Exception:
            pass
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return str(value.isoformat())
        except Exception:
            pass
    return str(value).strip()


def looks_like_xlsx_file(path: Path) -> bool:
    try:
        if not zipfile.is_zipfile(path):
            return False
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
        return "xl/workbook.xml" in names
    except Exception:
        return False


def looks_like_pptx_file(path: Path) -> bool:
    try:
        if not zipfile.is_zipfile(path):
            return False
        with zipfile.ZipFile(path, "r") as zf:
            names = set(zf.namelist())
        if "ppt/presentation.xml" not in names:
            return False
        return any(name.startswith("ppt/slides/slide") and name.endswith(".xml") for name in names)
    except Exception:
        return False


def _extract_xlsx(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook  # lazy import
    except Exception as exc:
        raise RuntimeError(
            "解析 .xlsx 需要依赖 openpyxl。请执行 `pip install -r requirements.txt` 后重试。"
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        lines: list[str] = ["[Excel 工作簿解析]"]
        total_chars = len(lines[0])
        truncated = False
        for sheet in wb.worksheets:
            title = (sheet.title or "").strip() or "Sheet"
            sheet_header = f"\n--- Sheet: {title} ---"
            lines.append(sheet_header)
            total_chars += len(sheet_header)
            if total_chars >= max_chars:
                truncated = True
                break

            sheet_rows = 0
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [_xlsx_cell_to_text(cell) for cell in row]
                while cells and not cells[-1]:
                    cells.pop()
                if not cells or not any(cells):
                    continue

                row_line = f"{row_idx}: " + " | ".join(cells)
                lines.append(row_line)
                total_chars += len(row_line)
                sheet_rows += 1
                if total_chars >= max_chars:
                    truncated = True
                    break

            if sheet_rows == 0:
                empty_line = "[空表或无可读内容]"
                lines.append(empty_line)
                total_chars += len(empty_line)
            if truncated:
                break

        if truncated:
            lines.append("\n[内容已截断，工作簿内容较大]")
        return truncate_text("\n".join(lines), max_chars)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _ppt_xml_to_lines(raw_xml: bytes, per_slide_limit: int = 40) -> list[str]:
    try:
        root = ET.fromstring(raw_xml)
    except Exception:
        return []
    lines: list[str] = []
    current: list[str] = []
    for node in root.iter():
        if _xml_local_name(node.tag) != "t":
            continue
        text = " ".join(str(node.text or "").split()).strip()
        if not text:
            continue
        current.append(text)
        # Keep nearby text runs together as one line.
        if len(current) >= 12:
            merged = " ".join(current).strip()
            if merged:
                lines.append(merged)
            current = []
        if len(lines) >= per_slide_limit:
            break
    if current and len(lines) < per_slide_limit:
        merged = " ".join(current).strip()
        if merged:
            lines.append(merged)
    # Deduplicate preserving order.
    out: list[str] = []
    seen: set[str] = set()
    for line in lines:
        key = line.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(line)
    return out


def _extract_pptx(path: Path, max_chars: int) -> str:
    if not zipfile.is_zipfile(path):
        return "[PPTX 解析失败: 文件不是合法 ZIP 容器]"

    with zipfile.ZipFile(path, "r") as zf:
        names = zf.namelist()
        slide_names = [
            name
            for name in names
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        ]
        if not slide_names:
            return "[PPTX 解析结果为空: 未找到 slide XML]"

        def sort_key(name: str) -> tuple[int, str]:
            m = re.search(r"slide(\d+)\.xml$", name)
            return (int(m.group(1)) if m else 10**9, name)

        slide_names.sort(key=sort_key)

        lines: list[str] = ["[PowerPoint 文档解析]"]
        lines.append(f"幻灯片数量: {len(slide_names)}")
        total = sum(len(x) for x in lines) + 1

        for idx, slide_name in enumerate(slide_names, start=1):
            try:
                raw_xml = zf.read(slide_name)
            except Exception:
                raw_xml = b""
            slide_lines = _ppt_xml_to_lines(raw_xml, per_slide_limit=36)
            header = f"\n--- Slide {idx} ---"
            if total + len(header) >= max_chars:
                lines.append("\n[内容已截断，幻灯片内容较多]")
                break
            lines.append(header)
            total += len(header)

            if not slide_lines:
                placeholder = "[未提取到文本（可能是纯图片页）]"
                if total + len(placeholder) >= max_chars:
                    lines.append("\n[内容已截断，幻灯片内容较多]")
                    break
                lines.append(placeholder)
                total += len(placeholder)
                continue

            for line in slide_lines:
                entry = f"- {line}"
                if total + len(entry) + 1 >= max_chars:
                    lines.append("\n[内容已截断，幻灯片内容较多]")
                    break
                lines.append(entry)
                total += len(entry) + 1
            else:
                continue
            break

    return truncate_text("\n".join(lines), max_chars)


def _html_to_text(html: str) -> str:
    raw = html or ""
    raw = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
    raw = re.sub(r"(?i)<br\\s*/?>", "\n", raw)
    raw = re.sub(r"(?i)</(p|div|li|tr|h1|h2|h3|h4|h5|h6|section|article)>", "\n", raw)
    raw = re.sub(r"(?s)<[^>]+>", " ", raw)
    raw = unescape(raw)
    lines: list[str] = []
    for line in raw.splitlines():
        normalized = re.sub(r"\s+", " ", line).strip()
        if normalized:
            lines.append(normalized)
    return "\n".join(lines)


def _xml_local_name(tag: str) -> str:
    raw = str(tag or "").strip()
    if "}" in raw:
        raw = raw.rsplit("}", 1)[-1]
    return raw.lower()


def _find_first_child_text(node: ET.Element, *names: str) -> str:
    wanted = {name.lower() for name in names if name}
    for child in list(node):
        if _xml_local_name(child.tag) not in wanted:
            continue
        text = "".join(child.itertext()).strip()
        if text:
            return text
    return ""


def _extract_xml_feed(path: Path, max_chars: int) -> str:
    raw_text = path.read_text(encoding="utf-8", errors="ignore")
    if not raw_text.strip():
        return ""

    try:
        root = ET.fromstring(raw_text)
    except Exception:
        return truncate_text(raw_text, max_chars)

    lines: list[str] = []
    root_name = _xml_local_name(root.tag)

    if root_name == "feed":
        lines.append("[Atom Feed 解析]")
        title = _find_first_child_text(root, "title")
        subtitle = _find_first_child_text(root, "subtitle", "tagline")
        updated = _find_first_child_text(root, "updated")
        feed_id = _find_first_child_text(root, "id")
        author_name = ""
        for child in list(root):
            if _xml_local_name(child.tag) == "author":
                author_name = _find_first_child_text(child, "name") or "".join(child.itertext()).strip()
                if author_name:
                    break

        if title:
            lines.append(f"标题: {title}")
        if subtitle:
            lines.append(f"副标题: {subtitle}")
        if updated:
            lines.append(f"更新时间: {updated}")
        if author_name:
            lines.append(f"作者: {author_name}")
        if feed_id:
            lines.append(f"Feed ID: {feed_id}")

        entries = [child for child in list(root) if _xml_local_name(child.tag) == "entry"]
        if entries:
            lines.append("条目:")
        for idx, entry in enumerate(entries, start=1):
            entry_title = _find_first_child_text(entry, "title") or f"entry_{idx}"
            entry_updated = _find_first_child_text(entry, "updated", "published")
            entry_summary = _find_first_child_text(entry, "summary", "content")
            lines.append(f"{idx}. {entry_title}")
            if entry_updated:
                lines.append(f"   更新时间: {entry_updated}")
            if entry_summary:
                entry_summary_clean = re.sub(r"\s+", " ", entry_summary).strip()
                lines.append(f"   摘要: {entry_summary_clean}")
    elif root_name == "rss":
        lines.append("[RSS Feed 解析]")
        channel = next((child for child in list(root) if _xml_local_name(child.tag) == "channel"), None)
        channel_node = channel or root
        title = _find_first_child_text(channel_node, "title")
        description = _find_first_child_text(channel_node, "description")
        updated = _find_first_child_text(channel_node, "lastbuilddate", "pubdate")
        if title:
            lines.append(f"标题: {title}")
        if description:
            lines.append(f"描述: {description}")
        if updated:
            lines.append(f"更新时间: {updated}")

        items = [child for child in list(channel_node) if _xml_local_name(child.tag) == "item"]
        if items:
            lines.append("条目:")
        for idx, item in enumerate(items, start=1):
            item_title = _find_first_child_text(item, "title") or f"item_{idx}"
            item_date = _find_first_child_text(item, "pubdate")
            item_desc = _find_first_child_text(item, "description", "summary")
            lines.append(f"{idx}. {item_title}")
            if item_date:
                lines.append(f"   时间: {item_date}")
            if item_desc:
                item_desc_clean = re.sub(r"\s+", " ", item_desc).strip()
                lines.append(f"   摘要: {item_desc_clean}")
    else:
        return truncate_text(raw_text, max_chars)

    return truncate_text("\n".join(lines).strip(), max_chars)


def _decode_bytes_best_effort(raw: bytes) -> str:
    if not raw:
        return ""
    for encoding in ("utf-8", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            out = raw.decode(encoding, errors="ignore")
        except Exception:
            continue
        if out.strip():
            return out
    return raw.decode("utf-8", errors="ignore")


def _looks_binaryish_text(text: str) -> bool:
    if not text:
        return False
    sample = text[:4096]
    if not sample:
        return False

    bad = 0
    for ch in sample:
        code = ord(ch)
        if code == 0:
            bad += 3
        elif code < 32 and ch not in "\n\r\t":
            bad += 1

    ratio = bad / max(1, len(sample))
    return ratio >= 0.02


def looks_like_outlook_msg_bytes(raw: bytes) -> bool:
    if not raw or not raw.startswith(_OLE2_MAGIC):
        return False
    head = raw[: max(4096, min(len(raw), 512 * 1024))]
    if any(marker in head for marker in _MSG_MARKERS_ASCII):
        return True
    if any(marker in head for marker in _MSG_MARKERS_UTF16):
        return True
    return False


def looks_like_outlook_msg_file(path: Path) -> bool:
    try:
        with path.open("rb") as fp:
            head = fp.read(512 * 1024)
    except Exception:
        return False
    return looks_like_outlook_msg_bytes(head)


def _extract_msg_body(msg: object) -> str:
    body = ""
    try:
        plain = getattr(msg, "body", None)
        if isinstance(plain, str):
            body = plain.strip()
        elif isinstance(plain, (bytes, bytearray)):
            body = _decode_bytes_best_effort(bytes(plain)).strip()
    except Exception:
        body = ""
    if body and not _looks_binaryish_text(body):
        return body

    try:
        html_body = getattr(msg, "htmlBody", None)
        if isinstance(html_body, (bytes, bytearray)):
            html_body = _decode_bytes_best_effort(bytes(html_body))
        if isinstance(html_body, str) and html_body.strip():
            html_text = _html_to_text(html_body).strip()
            if html_text and not _looks_binaryish_text(html_text):
                return html_text
    except Exception:
        pass

    try:
        rtf_body = getattr(msg, "rtfBody", None)
        deencap = getattr(msg, "deencapsulateBody", None)
        if rtf_body and callable(deencap):
            try:
                from extract_msg.enums import DeencapType  # lazy import

                plain_rtf = deencap(rtf_body, DeencapType.PLAIN)
            except Exception:
                plain_rtf = None
            if isinstance(plain_rtf, (bytes, bytearray)):
                plain_rtf = _decode_bytes_best_effort(bytes(plain_rtf))
            if isinstance(plain_rtf, str):
                plain_rtf = plain_rtf.strip()
                if plain_rtf and not _looks_binaryish_text(plain_rtf):
                    return plain_rtf
    except Exception:
        pass

    return ""


def _format_msg_attachment_line(att: object, idx: int) -> str:
    name = (
        (getattr(att, "longFilename", None) or "")
        or (getattr(att, "filename", None) or "")
        or (getattr(att, "name", None) or "")
        or f"attachment_{idx}"
    )
    extras: list[str] = []

    att_type = str(getattr(att, "type", "") or "").strip()
    if att_type:
        extras.append(att_type.split(".")[-1].lower())

    mime = (getattr(att, "mimetype", None) or "").strip()
    if mime:
        extras.append(mime)

    data = None
    try:
        data = getattr(att, "data", None)
    except Exception:
        data = None

    if isinstance(data, (bytes, bytearray)):
        extras.append(f"{len(data)} bytes")
    else:
        nested_subject = (getattr(data, "subject", None) or "").strip() if data is not None else ""
        if nested_subject:
            extras.append(f"嵌套邮件: {nested_subject}")

    if extras:
        return f"- {name} ({', '.join(extras)})"
    return f"- {name}"


def _extract_outlook_msg(path: Path, max_chars: int) -> str:
    try:
        import extract_msg  # lazy import
    except Exception as exc:
        raise RuntimeError(
            "解析 .msg 需要依赖 extract-msg。请执行 `pip install -r requirements.txt` 后重试。"
        ) from exc

    msg = extract_msg.openMsg(str(path), strict=False, delayAttachments=False)
    try:
        subject = (msg.subject or "").strip()
        sender = (msg.sender or "").strip()
        to = (msg.to or "").strip()
        cc = (msg.cc or "").strip()
        date = str(msg.date or "").strip()
        class_type = str(getattr(msg, "classType", "") or "").strip()
        body = _extract_msg_body(msg)

        attachment_lines: list[str] = []
        for idx, att in enumerate(getattr(msg, "attachments", []) or [], start=1):
            attachment_lines.append(_format_msg_attachment_line(att, idx))

        sections: list[str] = ["[Outlook MSG 邮件解析]"]
        if class_type:
            sections.append(f"消息类型: {class_type}")
        if subject:
            sections.append(f"主题: {subject}")
        if sender:
            sections.append(f"发件人: {sender}")
        if to:
            sections.append(f"收件人: {to}")
        if cc:
            sections.append(f"抄送: {cc}")
        if date:
            sections.append(f"时间: {date}")
        if attachment_lines:
            sections.append("附件列表:")
            sections.extend(attachment_lines)
        if body:
            sections.append("\n--- 正文 ---\n")
            sections.append(body)
        else:
            sections.append("\n--- 正文 ---\n")
            sections.append("[未提取到可读正文：该邮件可能仅包含附件、图片或受限富文本内容]")

        return truncate_text("\n".join(sections).strip(), max_chars)
    finally:
        close = getattr(msg, "close", None)
        if callable(close):
            try:
                close()
            except Exception:
                pass


def extract_document_text(path: str, max_chars: int) -> str | None:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    plain_suffixes = {
        ".atom",
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".log",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".html",
        ".css",
        ".yaml",
        ".yml",
        ".xml",
        ".rss",
    }

    try:
        if suffix in {".atom", ".rss", ".xml"}:
            return _extract_xml_feed(file_path, max_chars)
        if suffix in plain_suffixes:
            return _read_plain_text(file_path, max_chars)
        if suffix == ".pdf":
            return _extract_pdf(file_path, max_chars)
        if suffix == ".docx":
            return _extract_docx(file_path, max_chars)
        if suffix in _XLSX_SUFFIXES:
            return _extract_xlsx(file_path, max_chars)
        if suffix in _PPTX_SUFFIXES:
            return _extract_pptx(file_path, max_chars)
        if suffix == ".xls":
            return "[暂不支持 .xls（二进制 Excel）直接解析，请先另存为 .xlsx 后再读取]"
        if suffix == ".ppt":
            return "[暂不支持 .ppt（二进制 PowerPoint）直接解析，请先另存为 .pptx 后再读取]"
        if suffix in {".zip", ".bin"} and looks_like_xlsx_file(file_path):
            return _extract_xlsx(file_path, max_chars)
        if suffix in {".zip", ".bin"} and looks_like_pptx_file(file_path):
            return _extract_pptx(file_path, max_chars)
        if suffix == ".msg" or looks_like_outlook_msg_file(file_path):
            return _extract_outlook_msg(file_path, max_chars)
    except Exception as exc:
        return f"[文档解析失败: {exc}]"

    return None


def _heic_to_jpeg_bytes(path: Path) -> bytes:
    try:
        from PIL import Image
        from pillow_heif import register_heif_opener

        register_heif_opener()
        image = Image.open(path)
        rgb = image.convert("RGB")
        buffer = io.BytesIO()
        rgb.save(buffer, format="JPEG", quality=92)
        return buffer.getvalue()
    except Exception as exc:
        raise RuntimeError(
            "HEIC/HEIF conversion requires pillow-heif. Please install dependencies from requirements.txt."
        ) from exc


def image_to_data_url_with_meta(path: str, mime: str) -> tuple[str, str | None]:
    """
    Returns (data_url, warning). For HEIC, fallback to original HEIC payload
    when local conversion is unavailable, so capable gateways can still consume it.
    """
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    raw: bytes
    out_mime = mime
    warning: str | None = None

    is_heic = suffix in {".heic", ".heif"} or mime in {"image/heic", "image/heif"}
    if is_heic:
        try:
            raw = _heic_to_jpeg_bytes(file_path)
            out_mime = "image/jpeg"
        except Exception:
            raw = file_path.read_bytes()
            out_mime = mime if mime.startswith("image/") else "image/heic"
            warning = "HEIC 未本地转码，已原始上传；若网关不支持 HEIC，请先转 JPG/PNG。"
    else:
        raw = file_path.read_bytes()

    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{out_mime};base64,{encoded}", warning


def image_to_data_url(path: str, mime: str) -> str:
    data_url, _ = image_to_data_url_with_meta(path, mime)
    return data_url


def summarize_file_payload(path: str, max_bytes: int = 768, max_text_chars: int = 1200) -> str:
    file_path = Path(path)
    raw = file_path.read_bytes()
    head = raw[:max_bytes]

    if not head:
        return "[空文件]"

    text_bytes = b"\n\r\t\b\f" + bytes(range(32, 127))
    non_text = sum(1 for b in head if b not in text_bytes)
    is_binary = b"\x00" in head or (non_text / len(head)) > 0.30

    if not is_binary:
        text = head.decode("utf-8", errors="ignore")
        text = text[:max_text_chars]
        return f"[文本预览，文件大小 {len(raw)} bytes]\\n{text}"

    hex_preview = " ".join(f"{b:02x}" for b in head[:128])
    return (
        f"[二进制预览，文件大小 {len(raw)} bytes，前 {min(len(head),128)} bytes(hex)]\\n"
        f"{hex_preview}"
    )
